"""
Sistema de Ponto - Aplicação Web

Aplicação Flask responsável por:

- Autenticação de usuários
- Registro de ponto
- Cálculo de horas trabalhadas
- Persistência híbrida (SQLite local / PostgreSQL produção)

Compatível com:
- Render (PostgreSQL via DATABASE_URL)
- Execução local (SQLite)

O banco é selecionado automaticamente via variável de ambiente.
"""

import os
import secrets
import sqlite3
from urllib.parse import urlparse
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from typing import Optional, Tuple, Any, Dict
import threading
from io import BytesIO

import psycopg2
from flask import Flask, render_template, request, redirect, session, flash, jsonify, send_file, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


# ==========================================================
# CONFIGURAÇÃO DA APLICAÇÃO
# ==========================================================

app = Flask(__name__)

app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev_secret_key")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

DATABASE_PATH = os.getenv("DATABASE_PATH", "web/database.db")
DATABASE_URL = os.getenv("DATABASE_URL")
TIMEZONE = os.getenv("APP_TIMEZONE", "America/Sao_Paulo")

IS_POSTGRES = bool(DATABASE_URL)

_db_init_lock = threading.Lock()
_db_init_done = False


# ==========================================================
# BANCO DE DADOS
# ==========================================================

def conectar():
    """
    Retorna conexão com banco de dados.

    - Se DATABASE_URL estiver definida → PostgreSQL.
    - Caso contrário → SQLite local.
    """
    if IS_POSTGRES:
        result = urlparse(DATABASE_URL)
        return psycopg2.connect(
            dbname=result.path[1:],
            user=result.username,
            password=result.password,
            host=result.hostname,
            port=result.port,
        )
    return sqlite3.connect(DATABASE_PATH)


def sql(query: str) -> str:
    """
    Converte placeholders automaticamente.

    SQLite usa '?'.
    PostgreSQL usa '%s'.

    Essa função garante compatibilidade entre os dois.
    """
    if IS_POSTGRES:
        return query.replace("?", "%s")
    return query


def criar_banco() -> None:
    """
    Cria as tabelas necessárias caso não existam.
    Garante também compatibilidade com bancos legados:
    - Adiciona colunas ausentes na tabela `registros` (quando a estrutura antiga
      não possuía campos de almoço/saída final).
    """
    conn = conectar()
    cursor = conn.cursor()

    if IS_POSTGRES:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                nome_funcionario TEXT,
                nome_exibicao TEXT,
                nome_empresa TEXT,
                horas_diarias_esperadas_min INTEGER,
                role TEXT NOT NULL DEFAULT 'user'
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS registros (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES usuarios(id),
                data TEXT,
                entrada_manha TEXT,
                saida_almoco TEXT,
                volta_almoco TEXT,
                saida_final TEXT,
                corrigido_manual BOOLEAN NOT NULL DEFAULT FALSE,
                motivo_correcao TEXT,
                corrigido_em TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS invites (
                id SERIAL PRIMARY KEY,
                email TEXT NOT NULL,
                empresa TEXT NOT NULL,
                token TEXT UNIQUE NOT NULL,
                usado BOOLEAN NOT NULL DEFAULT FALSE,
                data_criacao TEXT NOT NULL
            )
        """)

        # Migração: adiciona colunas ausentes em bases antigas
        for coluna in ("entrada_manha", "saida_almoco", "volta_almoco", "saida_final"):
            cursor.execute(f"ALTER TABLE registros ADD COLUMN IF NOT EXISTS {coluna} TEXT")
        cursor.execute("ALTER TABLE registros ADD COLUMN IF NOT EXISTS corrigido_manual BOOLEAN NOT NULL DEFAULT FALSE")
        cursor.execute("ALTER TABLE registros ADD COLUMN IF NOT EXISTS motivo_correcao TEXT")
        cursor.execute("ALTER TABLE registros ADD COLUMN IF NOT EXISTS corrigido_em TEXT")

        # Migração: adiciona colunas do perfil em bases antigas
        cursor.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS nome_funcionario TEXT")
        cursor.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS nome_exibicao TEXT")
        cursor.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS nome_empresa TEXT")
        cursor.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS horas_diarias_esperadas_min INTEGER")
        cursor.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS role TEXT NOT NULL DEFAULT 'user'")
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                nome_funcionario TEXT,
                nome_exibicao TEXT,
                nome_empresa TEXT,
                horas_diarias_esperadas_min INTEGER,
                role TEXT NOT NULL DEFAULT 'user'
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS registros (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                data TEXT,
                entrada_manha TEXT,
                saida_almoco TEXT,
                volta_almoco TEXT,
                saida_final TEXT,
                corrigido_manual INTEGER NOT NULL DEFAULT 0,
                motivo_correcao TEXT,
                corrigido_em TEXT,
                FOREIGN KEY(user_id) REFERENCES usuarios(id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS invites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT NOT NULL,
                empresa TEXT NOT NULL,
                token TEXT UNIQUE NOT NULL,
                usado INTEGER NOT NULL DEFAULT 0,
                data_criacao TEXT NOT NULL
            )
        """)

        # Migração: adiciona colunas ausentes em bases antigas (SQLite)
        cursor.execute("PRAGMA table_info(registros)")
        colunas_existentes = {row[1] for row in cursor.fetchall()}
        for coluna, tipo in (
            ("entrada_manha", "TEXT"),
            ("saida_almoco", "TEXT"),
            ("volta_almoco", "TEXT"),
            ("saida_final", "TEXT"),
            ("corrigido_manual", "INTEGER NOT NULL DEFAULT 0"),
            ("motivo_correcao", "TEXT"),
            ("corrigido_em", "TEXT"),
        ):
            if coluna not in colunas_existentes:
                try:
                    cursor.execute(f"ALTER TABLE registros ADD COLUMN {coluna} {tipo}")
                except sqlite3.OperationalError as e:
                    # Se dois processos tentarem migrar ao mesmo tempo, o segundo pode
                    # receber "duplicate column name". Nesse caso, basta ignorar.
                    if "duplicate column name" not in str(e).lower():
                        raise

        # Migração: adiciona colunas do perfil em bases antigas (SQLite)
        cursor.execute("PRAGMA table_info(usuarios)")
        colunas_usuarios = {row[1] for row in cursor.fetchall()}
        for coluna, tipo in (
            ("nome_funcionario", "TEXT"),
            ("nome_exibicao", "TEXT"),
            ("nome_empresa", "TEXT"),
            ("horas_diarias_esperadas_min", "INTEGER"),
            ("role", "TEXT NOT NULL DEFAULT 'user'"),
        ):
            if coluna not in colunas_usuarios:
                try:
                    cursor.execute(f"ALTER TABLE usuarios ADD COLUMN {coluna} {tipo}")
                except sqlite3.OperationalError as e:
                    if "duplicate column name" not in str(e).lower():
                        raise

    conn.commit()
    conn.close()


# ==========================================================
# FUNÇÕES AUXILIARES
# ==========================================================

def parse_hora(hora: Optional[str]) -> Optional[datetime]:
    """
    Faz parse de strings de hora vindas do banco.

    Aceita formatos legados como:
    - HH:MM
    - HH:MM:SS
    - HH:MM:SS.ffffff
    """
    if not hora:
        return None

    for fmt in ("%H:%M:%S", "%H:%M", "%H:%M:%S.%f"):
        try:
            return datetime.strptime(hora, fmt)
        except ValueError:
            continue

    return None


def agora() -> datetime:
    """
    Retorna datetime atual considerando timezone configurado.
    """
    return datetime.now(ZoneInfo(TIMEZONE))


def usuario_logado() -> bool:
    """
    Verifica se existe usuário autenticado na sessão.
    """
    return "user_id" in session


def usuario_atual_id() -> Optional[int]:
    user_id = session.get("user_id")
    return int(user_id) if user_id is not None else None


def api_erro(mensagem: str, status_code: int = 400):
    return jsonify({"success": False, "error": mensagem}), status_code


def api_ok(payload: Optional[Dict[str, Any]] = None, status_code: int = 200):
    data: Dict[str, Any] = {"success": True}
    if payload:
        data.update(payload)
    return jsonify(data), status_code


def exigir_login_api():
    if not usuario_logado():
        return api_erro("Não autenticado", 401)
    return None


def flash_ok(msg: str) -> None:
    flash(msg, "success")


def flash_erro(msg: str) -> None:
    flash(msg, "error")


def flash_aviso(msg: str) -> None:
    flash(msg, "warning")


def flash_info(msg: str) -> None:
    flash(msg, "info")


def obter_perfil_usuario(user_id: int) -> Dict[str, Any]:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            SELECT nome_funcionario, nome_exibicao, nome_empresa, horas_diarias_esperadas_min
            FROM usuarios
            WHERE id=?
        """),
        (user_id,),
    )
    row = cursor.fetchone()
    conn.close()

    if not row:
        return {"nome_funcionario": None, "nome_exibicao": None, "nome_empresa": None, "horas_diarias_esperadas_min": None}

    return {"nome_funcionario": row[0], "nome_exibicao": row[1], "nome_empresa": row[2], "horas_diarias_esperadas_min": row[3]}


def obter_usuario(user_id: int) -> Optional[Dict[str, Any]]:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            SELECT id, username, nome_funcionario, nome_exibicao, nome_empresa, horas_diarias_esperadas_min, role
            FROM usuarios
            WHERE id=?
        """),
        (user_id,),
    )
    row = cursor.fetchone()
    conn.close()

    if not row:
        return None

    return {
        "id": row[0],
        "username": row[1],
        "nome_funcionario": row[2],
        "nome_exibicao": row[3],
        "nome_empresa": row[4],
        "horas_diarias_esperadas_min": row[5],
        "role": row[6] or "user",
    }


def usuario_eh_admin(user_id: Optional[int] = None) -> bool:
    alvo_id = user_id if user_id is not None else usuario_atual_id()
    if alvo_id is None:
        return False
    usuario = obter_usuario(alvo_id)
    return bool(usuario and usuario.get("role") == "admin")


def empresa_normalizada(usuario: Optional[Dict[str, Any]]) -> str:
    if not usuario:
        return ""
    return (usuario.get("nome_empresa") or "").strip()


def nome_usuario_para_exibicao(usuario: Dict[str, Any]) -> str:
    nome_exibicao = (usuario.get("nome_exibicao") or "").strip()
    nome_funcionario = (usuario.get("nome_funcionario") or "").strip()
    username = (usuario.get("username") or "").strip()
    return nome_exibicao or nome_funcionario or username or "não informado"


def admin_pode_acessar_funcionario(admin_id: int, funcionario_id: int) -> bool:
    admin = obter_usuario(admin_id)
    funcionario = obter_usuario(funcionario_id)
    if not admin or not funcionario:
        return False
    if admin.get("role") != "admin":
        return False
    empresa_admin = empresa_normalizada(admin)
    return bool(empresa_admin) and empresa_admin == empresa_normalizada(funcionario)


def admin_pode_promover_usuario(admin_id: int, usuario_id: int) -> bool:
    if admin_id == usuario_id:
        return False
    usuario = obter_usuario(usuario_id)
    return bool(usuario and usuario.get("role") == "user" and admin_pode_acessar_funcionario(admin_id, usuario_id))


def obter_registro(registro_id: int) -> Optional[Dict[str, Any]]:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            SELECT id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final,
                   corrigido_manual, motivo_correcao, corrigido_em
            FROM registros
            WHERE id=?
        """),
        (registro_id,),
    )
    row = cursor.fetchone()
    conn.close()

    if not row:
        return None

    return {
        "id": row[0],
        "user_id": row[1],
        "data": row[2],
        "entrada_manha": row[3],
        "saida_almoco": row[4],
        "volta_almoco": row[5],
        "saida_final": row[6],
        "corrigido_manual": bool(row[7]),
        "motivo_correcao": row[8],
        "corrigido_em": row[9],
    }


def usuario_pode_editar_registro_proprio(user_id: int, registro_id: int) -> bool:
    registro = obter_registro(registro_id)
    return bool(registro and registro["user_id"] == user_id)


def admin_pode_editar_registro(admin_id: int, registro_id: int) -> bool:
    registro = obter_registro(registro_id)
    return bool(registro and admin_pode_acessar_funcionario(admin_id, int(registro["user_id"])))


def primeiro_nome(nome: Optional[str]) -> Optional[str]:
    if not nome:
        return None
    partes = [p for p in nome.strip().split() if p]
    return partes[0] if partes else None


def perfil_para_json(perfil: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "nome_funcionario": perfil.get("nome_funcionario"),
        "nome_exibicao": perfil.get("nome_exibicao"),
        "nome_empresa": perfil.get("nome_empresa"),
        "horas_diarias_esperadas_min": perfil.get("horas_diarias_esperadas_min"),
    }


def atualizar_perfil_usuario(user_id: int, dados: Dict[str, Any]) -> Tuple[bool, Optional[str]]:
    nome_funcionario = (dados.get("nome_funcionario") or "").strip() or None
    nome_exibicao = (dados.get("nome_exibicao") or "").strip() or None
    horas_texto = dados.get("horas_diarias_esperadas") or ""

    perfil_atual = obter_perfil_usuario(user_id)
    empresa_atual = (perfil_atual.get("nome_empresa") or "").strip()
    nome_empresa = empresa_atual or None
    if not empresa_atual and "nome_empresa" in dados:
        nome_empresa = (dados.get("nome_empresa") or "").strip() or None
        if not nome_empresa:
            return False, "Informe a empresa para continuar."

    horas_min = parse_horas_esperadas_min(horas_texto)
    if str(horas_texto).strip() and horas_min is None:
        return False, "Formato inválido em horas diárias esperadas. Use 8, 8.5 ou 08:00."

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            UPDATE usuarios
            SET nome_funcionario=?, nome_exibicao=?, nome_empresa=?, horas_diarias_esperadas_min=?
            WHERE id=?
        """),
        (nome_funcionario, nome_exibicao, nome_empresa, horas_min, user_id),
    )
    conn.commit()
    conn.close()

    return True, None


def promover_usuario_para_admin(user_id: int) -> Optional[Dict[str, Any]]:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(sql("SELECT id, username, role FROM usuarios WHERE id=?"), (user_id,))
    antes = cursor.fetchone()
    if not antes:
        conn.close()
        return None

    cursor.execute(sql("UPDATE usuarios SET role='admin' WHERE id=?"), (user_id,))
    conn.commit()
    cursor.execute(sql("SELECT id, username, role FROM usuarios WHERE id=?"), (user_id,))
    depois = cursor.fetchone()
    conn.close()

    if not depois:
        return None

    return {
        "id": depois[0],
        "username": depois[1],
        "role_anterior": antes[2] or "user",
        "role": depois[2] or "user",
    }


def obter_convite_por_token(token: str, apenas_pendente: bool = True) -> Optional[Dict[str, Any]]:
    token = (token or "").strip()
    if not token:
        return None

    conn = conectar()
    cursor = conn.cursor()
    query = "SELECT id, email, empresa, token, usado, data_criacao FROM invites WHERE token=?"
    if apenas_pendente:
        query += " AND usado=FALSE"
    cursor.execute(sql(query), (token,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return None

    return {
        "id": row[0],
        "email": row[1],
        "empresa": row[2],
        "token": row[3],
        "usado": bool(row[4]),
        "data_criacao": row[5],
    }


def criar_convite(email: str, empresa: str) -> Dict[str, Any]:
    token = secrets.token_urlsafe(24)
    data_criacao = agora().isoformat(timespec="seconds")

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            INSERT INTO invites (email, empresa, token, usado, data_criacao)
            VALUES (?, ?, ?, FALSE, ?)
        """),
        (email, empresa, token, data_criacao),
    )
    conn.commit()
    conn.close()

    return {"email": email, "empresa": empresa, "token": token, "data_criacao": data_criacao}


def listar_convites_pendentes(empresa: str) -> list[Dict[str, Any]]:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            SELECT id, email, empresa, token, data_criacao
            FROM invites
            WHERE empresa=? AND usado=FALSE
            ORDER BY id DESC
        """),
        (empresa,),
    )
    rows = cursor.fetchall()
    conn.close()
    return [
        {"id": row[0], "email": row[1], "empresa": row[2], "token": row[3], "data_criacao": row[4]}
        for row in rows
    ]


def marcar_convite_usado(convite_id: int) -> None:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(sql("UPDATE invites SET usado=TRUE WHERE id=?"), (convite_id,))
    conn.commit()
    conn.close()


def cancelar_convite_empresa(convite_id: int, empresa: str) -> bool:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("UPDATE invites SET usado=TRUE WHERE id=? AND empresa=? AND usado=FALSE"),
        (convite_id, empresa),
    )
    alterados = cursor.rowcount
    conn.commit()
    conn.close()
    return bool(alterados)


def excluir_registro_usuario(user_id: int, registro_id: int) -> bool:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(sql("DELETE FROM registros WHERE id=? AND user_id=?"), (registro_id, user_id))
    apagados = cursor.rowcount
    conn.commit()
    conn.close()
    return bool(apagados)


def excluir_conta_usuario(user_id: int) -> None:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(sql("DELETE FROM registros WHERE user_id=?"), (user_id,))
    cursor.execute(sql("DELETE FROM usuarios WHERE id=?"), (user_id,))
    conn.commit()
    conn.close()


def registrar_ponto(user_id: int, quando: datetime) -> Dict[str, Any]:
    hoje = quando.strftime("%Y-%m-%d")
    hora_atual = quando.strftime("%H:%M:%S")

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            SELECT id, entrada_manha, saida_almoco, volta_almoco, saida_final
            FROM registros
            WHERE user_id=? AND data=?
            ORDER BY id DESC LIMIT 1
        """),
        (user_id, hoje),
    )
    registro = cursor.fetchone()

    acao = "criado"
    campo = "entrada_manha"
    registro_id = None

    if not registro:
        cursor.execute(
            sql("INSERT INTO registros (user_id, data, entrada_manha) VALUES (?, ?, ?)"),
            (user_id, hoje, hora_atual),
        )
        registro_id = cursor.lastrowid if not IS_POSTGRES else None
    else:
        registro_id = registro[0]
        saida_almoco = registro[2]
        volta_almoco = registro[3]
        saida_final = registro[4]

        if not saida_almoco:
            cursor.execute(sql("UPDATE registros SET saida_almoco=? WHERE id=?"), (hora_atual, registro_id))
            acao = "atualizado"
            campo = "saida_almoco"
        elif not volta_almoco:
            cursor.execute(sql("UPDATE registros SET volta_almoco=? WHERE id=?"), (hora_atual, registro_id))
            acao = "atualizado"
            campo = "volta_almoco"
        elif not saida_final:
            cursor.execute(sql("UPDATE registros SET saida_final=? WHERE id=?"), (hora_atual, registro_id))
            acao = "atualizado"
            campo = "saida_final"
        else:
            cursor.execute(
                sql("INSERT INTO registros (user_id, data, entrada_manha) VALUES (?, ?, ?)"),
                (user_id, hoje, hora_atual),
            )
            acao = "criado"
            campo = "entrada_manha"
            registro_id = cursor.lastrowid if not IS_POSTGRES else None

    conn.commit()
    conn.close()

    return {"acao": acao, "campo": campo, "data": hoje, "hora": hora_atual, "registro_id": registro_id}


def parse_horas_esperadas_min(valor: Optional[str]) -> Optional[int]:
    """
    Aceita entrada do perfil como:
    - "8" ou "8.5" (horas)
    - "08:00" ou "8:30" (HH:MM)
    Retorna minutos (int) ou None.
    """
    if valor is None:
        return None

    texto = valor.strip()
    if not texto:
        return None

    if ":" in texto:
        partes = texto.split(":")
        if len(partes) != 2:
            return None
        try:
            horas = int(partes[0])
            minutos = int(partes[1])
        except ValueError:
            return None
        if horas < 0 or minutos < 0 or minutos >= 60:
            return None
        return horas * 60 + minutos

    try:
        horas_float = float(texto.replace(",", "."))
    except ValueError:
        return None
    if horas_float < 0:
        return None
    return int(round(horas_float * 60))


def formatar_horas_minutos(delta: timedelta) -> str:
    total_min = int(round(delta.total_seconds() / 60))
    sinal = "+" if total_min >= 0 else "-"
    total_min_abs = abs(total_min)
    horas = total_min_abs // 60
    minutos = total_min_abs % 60
    return f"{sinal}{horas}h {minutos:02d}m"


def formatar_banco_horas(delta: timedelta) -> str:
    if int(delta.total_seconds()) == 0:
        return "0h 00m"
    return formatar_horas_minutos(delta)


def limites_mes_atual_iso() -> Tuple[str, str]:
    """
    Retorna (primeiro_dia, ultimo_dia) do mês atual em formato ISO YYYY-MM-DD,
    baseado no timezone configurado pela função `agora()`.
    """
    hoje_local = agora().date()
    primeiro = hoje_local.replace(day=1)
    if primeiro.month == 12:
        proximo = primeiro.replace(year=primeiro.year + 1, month=1, day=1)
    else:
        proximo = primeiro.replace(month=primeiro.month + 1, day=1)
    ultimo = proximo - timedelta(days=1)
    return primeiro.isoformat(), ultimo.isoformat()


def formatar_media_diaria(total: timedelta, dias: int) -> str:
    if dias <= 0:
        return "0h 00m"
    media_seg = total.total_seconds() / dias
    return formatar_duracao_sem_sinal(timedelta(seconds=media_seg))


def calcular_saldo_dia(total_trabalhado: timedelta, esperado_min: Optional[int]) -> Optional[timedelta]:
    if esperado_min is None:
        return None
    return total_trabalhado - timedelta(minutes=int(esperado_min))


def horas_decimal(delta: timedelta) -> float:
    return round(delta.total_seconds() / 3600, 2)


def formatar_duracao_sem_sinal(delta: timedelta) -> str:
    total_min = int(round(delta.total_seconds() / 60))
    total_min_abs = abs(total_min)
    horas = total_min_abs // 60
    minutos = total_min_abs % 60
    return f"{horas}h {minutos:02d}m"


def validar_periodo(data_inicio: Optional[str], data_fim: Optional[str]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Valida datas no formato ISO (YYYY-MM-DD) e garante que data_inicio <= data_fim.
    Retorna (inicio, fim, erro). inicio/fim são strings ISO normalizadas (ou None).
    """
    inicio = (data_inicio or "").strip() or None
    fim = (data_fim or "").strip() or None

    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date() if inicio else None
    except ValueError:
        return None, None, "Data inicial inválida."

    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date() if fim else None
    except ValueError:
        return None, None, "Data final inválida."

    if inicio_dt and fim_dt and inicio_dt > fim_dt:
        return inicio, fim, "Data inicial não pode ser maior que a data final."

    return (inicio_dt.isoformat() if inicio_dt else None), (fim_dt.isoformat() if fim_dt else None), None


def normalizar_hora_formulario(valor: Optional[str], nome_campo: str) -> Tuple[Optional[str], Optional[str]]:
    texto = (valor or "").strip()
    if not texto:
        return None, None

    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            hora = datetime.strptime(texto, fmt)
            return hora.strftime("%H:%M"), None
        except ValueError:
            continue

    return None, f"{nome_campo} inválido. Use o formato HH:MM."


def dados_correcao_do_formulario() -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    data_texto = (request.form.get("data") or "").strip()
    motivo = (request.form.get("motivo_correcao") or "").strip()

    if not motivo:
        return None, "Informe o motivo da correção."

    try:
        data = datetime.strptime(data_texto, "%Y-%m-%d").date().isoformat()
    except ValueError:
        return None, "Data inválida."

    campos_hora = (
        ("entrada_manha", "Entrada"),
        ("saida_almoco", "Saída almoço"),
        ("volta_almoco", "Volta almoço"),
        ("saida_final", "Saída final"),
    )
    dados: Dict[str, Any] = {"data": data, "motivo_correcao": motivo}
    for campo, rotulo in campos_hora:
        hora, erro = normalizar_hora_formulario(request.form.get(campo), rotulo)
        if erro:
            return None, erro
        dados[campo] = hora

    return dados, None


def atualizar_registro_manual(registro_id: int, dados: Dict[str, Any]) -> bool:
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            UPDATE registros
            SET data=?,
                entrada_manha=?,
                saida_almoco=?,
                volta_almoco=?,
                saida_final=?,
                corrigido_manual=TRUE,
                motivo_correcao=?,
                corrigido_em=?
            WHERE id=?
        """),
        (
            dados["data"],
            dados["entrada_manha"],
            dados["saida_almoco"],
            dados["volta_almoco"],
            dados["saida_final"],
            dados["motivo_correcao"],
            agora().isoformat(timespec="seconds"),
            registro_id,
        ),
    )
    alterados = cursor.rowcount
    conn.commit()
    conn.close()
    return bool(alterados)


def aplicar_filtro_periodo_sql(sql_base: str, inicio: Optional[str], fim: Optional[str]) -> Tuple[str, Tuple[Any, ...]]:
    """
    Retorna (sql_filtrado, params_adicionais) para filtrar por data.
    As datas são strings ISO (YYYY-MM-DD), compatíveis com comparação lexicográfica.
    """
    if inicio and fim:
        return sql_base + "\n            AND data BETWEEN ? AND ?", (inicio, fim)
    if inicio:
        return sql_base + "\n            AND data >= ?", (inicio,)
    if fim:
        return sql_base + "\n            AND data <= ?", (fim,)
    return sql_base, ()


def calcular_total_registro(registro: Tuple[Any, ...]) -> timedelta:
    """
    Calcula o total trabalhado de um registro.

    Recebe tupla retornada do banco.
    Retorna timedelta correspondente ao total do dia.
    """
    entrada, saida_almoco, volta_almoco, saida_final = (
        registro[3],
        registro[4],
        registro[5],
        registro[6],
    )

    total = timedelta()

    entrada_dt = parse_hora(entrada)
    saida_almoco_dt = parse_hora(saida_almoco)
    volta_almoco_dt = parse_hora(volta_almoco)
    saida_final_dt = parse_hora(saida_final)

    if entrada_dt and saida_almoco_dt and saida_almoco_dt >= entrada_dt:
        total += (saida_almoco_dt - entrada_dt)

    if volta_almoco_dt and saida_final_dt and saida_final_dt >= volta_almoco_dt:
        total += (saida_final_dt - volta_almoco_dt)

    return total


def buscar_registros_usuario(user_id: int, data_inicio: Optional[str] = None, data_fim: Optional[str] = None) -> list[Tuple[Any, ...]]:
    conn = conectar()
    cursor = conn.cursor()
    sql_query, params_extra = aplicar_filtro_periodo_sql(
        """
            SELECT id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final,
                   corrigido_manual, motivo_correcao, corrigido_em
            FROM registros
            WHERE user_id=?
        """,
        data_inicio,
        data_fim,
    )
    sql_query += "\n            ORDER BY data DESC, id DESC"
    cursor.execute(sql(sql_query), (user_id,) + params_extra)
    registros = cursor.fetchall()
    conn.close()
    return registros


def calcular_banco_registros(registros_db: list[Tuple[Any, ...]], esperado_min: int) -> timedelta:
    banco = timedelta()
    for registro in registros_db:
        if not parse_hora(registro[6]):
            continue
        saldo = calcular_saldo_dia(calcular_total_registro(registro), esperado_min)
        if saldo is not None:
            banco += saldo
    return banco


def montar_registros_para_tabela(registros_db: list[Tuple[Any, ...]], esperado_min: int) -> list[Dict[str, Any]]:
    registros = []
    for registro in registros_db:
        total_linha = calcular_total_registro(registro)
        em_aberto = not bool(parse_hora(registro[6]))
        saldo_delta = None if em_aberto else calcular_saldo_dia(total_linha, esperado_min)
        saldo_str = "em aberto" if em_aberto else (formatar_horas_minutos(saldo_delta) if saldo_delta is not None else "-")
        saldo_css = ""
        if em_aberto:
            saldo_css = "text-muted"
        elif saldo_delta is not None:
            saldo_min = int(round(saldo_delta.total_seconds() / 60))
            if saldo_min > 0:
                saldo_css = "text-success"
            elif saldo_min < 0:
                saldo_css = "text-danger"
            else:
                saldo_css = "text-secondary"

        registros.append(
            {
                "id": registro[0],
                "data": registro[2],
                "entrada": registro[3],
                "saida_almoco": registro[4],
                "volta_almoco": registro[5],
                "saida_final": registro[6],
                "total": total_linha,
                "saldo": saldo_str,
                "saldo_css": saldo_css,
                "corrigido_manual": bool(registro[7]) if len(registro) > 7 else False,
                "motivo_correcao": registro[8] if len(registro) > 8 else None,
                "corrigido_em": registro[9] if len(registro) > 9 else None,
            }
        )
    return registros


def periodo_para_texto(data_inicio: Optional[str], data_fim: Optional[str], prefixo: str = "Período") -> str:
    if data_inicio and data_fim:
        return f"{prefixo}: {datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    if data_inicio:
        return f"{prefixo}: a partir de {datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    if data_fim:
        return f"{prefixo}: até {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    return f"{prefixo}: todos os registros"


# ==========================================================
# ROTAS
# ==========================================================

@app.before_request
def garantir_banco():
    global _db_init_done
    if _db_init_done:
        return

    with _db_init_lock:
        if _db_init_done:
            return
        criar_banco()
        _db_init_done = True


@app.context_processor
def contexto_usuario():
    return {"usuario_admin": usuario_eh_admin()}


@app.route("/", methods=["GET", "POST"])
def login():
    """
    Realiza autenticação do usuário.
    """
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if not username or not password:
            return render_template("login.html", erro="Preencha todos os campos")

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute(
            sql("SELECT * FROM usuarios WHERE username=?"),
            (username,)
        )
        user = cursor.fetchone()
        conn.close()

        if user and check_password_hash(user[2], password):
            session["user_id"] = user[0]
            return redirect("/dashboard")

        return render_template("login.html", erro="Usuário ou senha inválidos")

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    """
    Registra novo usuário.
    """
    token_convite = (request.values.get("token") or "").strip()
    convite = obter_convite_por_token(token_convite) if token_convite else None
    if token_convite and not convite:
        return render_template(
            "register.html",
            erro="Convite inválido ou já utilizado.",
            convite_bloqueado=True,
        )

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if not username or not password:
            return render_template("register.html", erro="Preencha todos os campos", convite=convite, token=token_convite)

        senha_hash = generate_password_hash(password)

        conn = conectar()
        cursor = conn.cursor()

        try:
            if convite:
                cursor.execute(
                    sql("INSERT INTO usuarios (username, password, nome_empresa) VALUES (?, ?, ?)"),
                    (username, senha_hash, convite["empresa"]),
                )
            else:
                cursor.execute(
                    sql("INSERT INTO usuarios (username, password) VALUES (?, ?)"),
                    (username, senha_hash),
                )
            conn.commit()
        except Exception:
            conn.close()
            return render_template("register.html", erro="Usuário já existe", convite=convite, token=token_convite)

        conn.close()
        if convite:
            marcar_convite_usado(convite["id"])
        return redirect("/")

    return render_template("register.html", convite=convite, token=token_convite)


@app.route("/dashboard")
def dashboard():
    """
    Exibe histórico de registros e total do dia atual.
    """
    if not usuario_logado():
        return redirect("/")

    data_inicio_raw = request.args.get("data_inicio")
    data_fim_raw = request.args.get("data_fim")
    data_inicio, data_fim, erro_periodo = validar_periodo(data_inicio_raw, data_fim_raw)
    if erro_periodo:
        flash_erro(erro_periodo)
        qs = ""
        if data_inicio_raw or data_fim_raw:
            qs = f"?data_inicio={data_inicio_raw or ''}&data_fim={data_fim_raw or ''}"
        return redirect("/dashboard" + qs)

    hoje = agora().strftime("%Y-%m-%d")

    conn = conectar()
    cursor = conn.cursor()

    perfil = obter_perfil_usuario(session["user_id"])
    esperado_min = perfil.get("horas_diarias_esperadas_min")
    if esperado_min is None:
        esperado_min = 8 * 60
    nome_exibicao = (perfil.get("nome_exibicao") or "").strip()
    nome_funcionario = (perfil.get("nome_funcionario") or "").strip()
    nome_dashboard = (
        nome_exibicao
        or (primeiro_nome(nome_funcionario) or "")
        or "não informado"
    )

    sql_query, params_extra = aplicar_filtro_periodo_sql(
        """
            SELECT id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final,
                   corrigido_manual, motivo_correcao, corrigido_em
            FROM registros
            WHERE user_id=?
        """,
        data_inicio,
        data_fim,
    )
    sql_query += "\n            ORDER BY data DESC, id DESC"

    cursor.execute(
        sql(sql_query),
        (session["user_id"],) + params_extra,
    )

    registros_db = cursor.fetchall()
    conn.close()

    registros = []
    total_hoje = timedelta()
    graf_labels: list[str] = []
    graf_saldo: list[float] = []
    banco_periodo = timedelta()

    mes_inicio, mes_fim = limites_mes_atual_iso()
    conn_mes = conectar()
    cursor_mes = conn_mes.cursor()
    cursor_mes.execute(
        sql("""
            SELECT id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final,
                   corrigido_manual, motivo_correcao, corrigido_em
            FROM registros
            WHERE user_id=? AND data BETWEEN ? AND ?
            ORDER BY data DESC, id DESC
        """),
        (session["user_id"], mes_inicio, mes_fim),
    )
    registros_mes = cursor_mes.fetchall()
    conn_mes.close()
    banco_mes = timedelta()
    total_mes_finalizado = timedelta()
    dias_trabalhados_set: set[str] = set()
    for registro_mes in registros_mes:
        total_mes = calcular_total_registro(registro_mes)
        em_aberto_mes = not bool(parse_hora(registro_mes[6]))
        if em_aberto_mes:
            continue
        saldo_mes = calcular_saldo_dia(total_mes, esperado_min)
        if saldo_mes is not None:
            banco_mes += saldo_mes
        total_mes_finalizado += total_mes
        dias_trabalhados_set.add(registro_mes[2])

    dias_trabalhados_mes = len(dias_trabalhados_set)
    media_diaria_mes = formatar_media_diaria(total_mes_finalizado, dias_trabalhados_mes)

    for registro in registros_db:
        total_linha = calcular_total_registro(registro)
        if registro[2] == hoje:
            total_hoje += total_linha

        saida_final = registro[6]
        em_aberto = not bool(parse_hora(saida_final))
        saldo_delta = None if em_aberto else calcular_saldo_dia(total_linha, esperado_min)
        saldo_str = "em aberto" if em_aberto else (formatar_horas_minutos(saldo_delta) if saldo_delta is not None else "-")
        saldo_css = ""
        if em_aberto:
            saldo_css = "text-muted"
        elif saldo_delta is not None:
            saldo_min = int(round(saldo_delta.total_seconds() / 60))
            if saldo_min > 0:
                saldo_css = "text-success"
            elif saldo_min < 0:
                saldo_css = "text-danger"
            else:
                saldo_css = "text-secondary"

        if not em_aberto:
            graf_labels.append(registro[2])
            graf_saldo.append(horas_decimal(saldo_delta or timedelta()))
            if saldo_delta is not None:
                banco_periodo += saldo_delta

        registros.append({
            "id": registro[0],
            "data": registro[2],
            "entrada": registro[3],
            "saida_almoco": registro[4],
            "volta_almoco": registro[5],
            "saida_final": registro[6],
            "total": total_linha,
            "esperado": timedelta(minutes=int(esperado_min)) if esperado_min is not None else None,
            "saldo": saldo_str,
            "saldo_css": saldo_css,
            "corrigido_manual": bool(registro[7]) if len(registro) > 7 else False,
            "motivo_correcao": registro[8] if len(registro) > 8 else None,
            "corrigido_em": registro[9] if len(registro) > 9 else None,
        })

    # Status do dia (hoje) — baseado no total do dia e se existe registro em aberto
    status_texto = "Não iniciada"
    status_classe = "secondary"
    saldo_hoje_str = "-"
    total_hoje_str = formatar_duracao_sem_sinal(total_hoje)

    registros_hoje = [r for r in registros_db if r[2] == hoje and r[3]]
    if registros_hoje:
        existe_em_aberto = any(not r[6] for r in registros_hoje)
        if existe_em_aberto:
            status_texto = "Em andamento"
            status_classe = "info"
        else:
            saldo_delta_hoje = calcular_saldo_dia(total_hoje, esperado_min)
            saldo_hoje_str = formatar_horas_minutos(saldo_delta_hoje or timedelta())
            if total_hoje >= timedelta(minutes=int(esperado_min)):
                status_texto = "Jornada completa"
                status_classe = "success"
            else:
                status_texto = "Jornada incompleta"
                status_classe = "warning"

    # Texto do botão de ponto (sem alterar a lógica da rota /bater)
    # Ordem do registro: (id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
    botao_ponto_texto = "Registrar entrada"
    if registros_hoje:
        # Pega o registro mais recente do dia para decidir o próximo passo
        ultimo = sorted(registros_hoje, key=lambda r: r[0], reverse=True)[0]
        entrada_manha = ultimo[3]
        saida_almoco = ultimo[4]
        volta_almoco = ultimo[5]
        saida_final = ultimo[6]

        if not entrada_manha:
            botao_ponto_texto = "Registrar entrada"
        elif not saida_almoco:
            botao_ponto_texto = "Registrar saída almoço"
        elif not volta_almoco:
            botao_ponto_texto = "Registrar volta almoço"
        elif not saida_final:
            botao_ponto_texto = "Registrar saída final"
        else:
            botao_ponto_texto = "Registrar novo ponto"

    return render_template(
        "dashboard.html",
        registros=registros,
        total_hoje=total_hoje,
        nome_funcionario=nome_dashboard,
        graf_labels=graf_labels,
        graf_saldo=graf_saldo,
        data_inicio=data_inicio or "",
        data_fim=data_fim or "",
        banco_periodo=formatar_horas_minutos(banco_periodo),
        banco_mes_atual=formatar_banco_horas(banco_mes),
        resumo_mes={
            "dias_trabalhados": dias_trabalhados_mes,
            "total_horas": formatar_duracao_sem_sinal(total_mes_finalizado),
            "media_diaria": media_diaria_mes,
            "banco_horas": formatar_banco_horas(banco_mes),
        },
        status_dia_texto=status_texto,
        status_dia_classe=status_classe,
        total_hoje_status=total_hoje_str,
        saldo_hoje_status=saldo_hoje_str,
        botao_ponto_texto=botao_ponto_texto,
    )


@app.route("/admin")
def admin_dashboard():
    if not usuario_logado():
        return redirect("/")
    admin_id = session["user_id"]
    admin = obter_usuario(admin_id)
    if not admin or admin.get("role") != "admin":
        flash_erro("Acesso restrito a administradores.")
        return redirect("/dashboard")

    empresa_admin = empresa_normalizada(admin)
    if not empresa_admin:
        return render_template(
            "admin.html",
            funcionarios=[],
            empresa_admin="Empresa não definida",
            aviso_admin="Defina sua empresa no Perfil para usar o painel admin.",
            convites=[],
        )

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            SELECT id, username, nome_funcionario, nome_exibicao, nome_empresa, horas_diarias_esperadas_min, role
            FROM usuarios
            WHERE COALESCE(nome_empresa, '')=?
            ORDER BY COALESCE(nome_funcionario, username), username
        """),
        (empresa_admin,),
    )
    funcionarios_db = cursor.fetchall()
    conn.close()

    mes_inicio, mes_fim = limites_mes_atual_iso()
    funcionarios = []
    for row in funcionarios_db:
        funcionario = {
            "id": row[0],
            "username": row[1],
            "nome_funcionario": row[2],
            "nome_exibicao": row[3],
            "nome_empresa": row[4],
            "horas_diarias_esperadas_min": row[5],
            "role": row[6] or "user",
        }
        esperado_min = funcionario.get("horas_diarias_esperadas_min") or 8 * 60
        registros_mes = buscar_registros_usuario(funcionario["id"], mes_inicio, mes_fim)
        todos_registros = buscar_registros_usuario(funcionario["id"])
        funcionarios.append(
            {
                "id": funcionario["id"],
                "nome": nome_usuario_para_exibicao(funcionario),
                "empresa": empresa_normalizada(funcionario) or "-",
                "role": funcionario["role"],
                "is_admin": funcionario["role"] == "admin",
                "pode_promover": admin_pode_promover_usuario(admin_id, funcionario["id"]),
                "total_registros": len(todos_registros),
                "banco_mes_atual": formatar_banco_horas(calcular_banco_registros(registros_mes, esperado_min)),
            }
        )

    return render_template(
        "admin.html",
        funcionarios=funcionarios,
        empresa_admin=empresa_admin or "-",
        aviso_admin=None,
        convites=listar_convites_pendentes(empresa_admin),
    )


@app.route("/admin/convites", methods=["POST"])
def admin_criar_convite():
    if not usuario_logado():
        return redirect("/")

    admin_id = session["user_id"]
    admin = obter_usuario(admin_id)
    if not admin or admin.get("role") != "admin":
        flash_erro("Acesso restrito a administradores.")
        return redirect("/dashboard")

    empresa_admin = empresa_normalizada(admin)
    if not empresa_admin:
        flash_erro("Defina sua empresa no Perfil para convidar usuários.")
        return redirect("/admin")

    email = (request.form.get("email") or "").strip().lower()
    if not email or "@" not in email:
        flash_erro("Informe um email válido para o convite.")
        return redirect("/admin")

    convite = criar_convite(email, empresa_admin)
    flash_info(f"Link de convite: /register?token={convite['token']}")
    return redirect("/admin")


@app.route("/admin/convites/<int:convite_id>/cancelar", methods=["POST"])
def admin_cancelar_convite(convite_id: int):
    if not usuario_logado():
        return redirect("/")

    admin_id = session["user_id"]
    admin = obter_usuario(admin_id)
    if not admin or admin.get("role") != "admin":
        flash_erro("Acesso restrito a administradores.")
        return redirect("/dashboard")

    empresa_admin = empresa_normalizada(admin)
    if not empresa_admin or not cancelar_convite_empresa(convite_id, empresa_admin):
        flash_erro("Convite não encontrado para esta empresa.")
        return redirect("/admin")

    flash_ok("Convite cancelado com sucesso.")
    return redirect("/admin")


@app.route("/setup-admin")
def setup_admin_temporario():
    if not usuario_logado():
        return redirect("/")

    if os.getenv("ALLOW_ADMIN_SETUP", "").lower() != "true":
        flash_erro("Setup de admin não autorizado.")
        return redirect("/dashboard")

    user_id = session["user_id"]
    usuario = promover_usuario_para_admin(user_id)
    if not usuario:
        flash_erro("Usuário não encontrado.")
        return redirect("/dashboard")

    print(
        "PROMOCAO_ADMIN "
        f"id={usuario['id']} username={usuario['username']} "
        f"role_anterior={usuario['role_anterior']} role={usuario['role']}",
        flush=True,
    )
    flash_ok("Usuário promovido a admin com sucesso")
    return redirect("/admin")


@app.route("/admin/usuarios/<int:usuario_id>/promover-admin", methods=["POST"])
def admin_promover_usuario(usuario_id: int):
    if not usuario_logado():
        return redirect("/")

    admin_id = session["user_id"]
    if not usuario_eh_admin(admin_id):
        flash_erro("Acesso restrito a administradores.")
        return redirect("/dashboard")

    if not admin_pode_promover_usuario(admin_id, usuario_id):
        flash_erro("Usuário não encontrado para esta empresa ou já é admin.")
        return redirect("/admin")

    usuario = promover_usuario_para_admin(usuario_id)
    if not usuario:
        flash_erro("Usuário não encontrado.")
        return redirect("/admin")

    flash_ok("Usuário promovido a admin com sucesso")
    return redirect("/admin")


@app.route("/admin/funcionarios/<int:funcionario_id>")
def admin_funcionario(funcionario_id: int):
    if not usuario_logado():
        return redirect("/")
    admin_id = session["user_id"]
    if not admin_pode_acessar_funcionario(admin_id, funcionario_id):
        flash_erro("Funcionário não encontrado para esta empresa.")
        return redirect("/admin" if usuario_eh_admin(admin_id) else "/dashboard")

    data_inicio_raw = request.args.get("data_inicio")
    data_fim_raw = request.args.get("data_fim")
    data_inicio, data_fim, erro_periodo = validar_periodo(data_inicio_raw, data_fim_raw)
    if erro_periodo:
        flash_erro(erro_periodo)
        qs = ""
        if data_inicio_raw or data_fim_raw:
            qs = f"?data_inicio={data_inicio_raw or ''}&data_fim={data_fim_raw or ''}"
        return redirect(f"/admin/funcionarios/{funcionario_id}" + qs)

    funcionario = obter_usuario(funcionario_id)
    if not funcionario:
        flash_erro("Funcionário não encontrado.")
        return redirect("/admin")

    esperado_min = funcionario.get("horas_diarias_esperadas_min") or 8 * 60
    registros_db = buscar_registros_usuario(funcionario_id, data_inicio, data_fim)
    registros = montar_registros_para_tabela(registros_db, esperado_min)

    return render_template(
        "admin_funcionario.html",
        funcionario=funcionario,
        funcionario_nome=nome_usuario_para_exibicao(funcionario),
        registros=registros,
        data_inicio=data_inicio or "",
        data_fim=data_fim or "",
        banco_periodo=formatar_banco_horas(calcular_banco_registros(registros_db, esperado_min)),
    )


@app.route("/admin/funcionarios/<int:funcionario_id>/export/excel")
def admin_export_excel(funcionario_id: int):
    if not usuario_logado():
        return redirect("/")
    admin_id = session["user_id"]
    if not admin_pode_acessar_funcionario(admin_id, funcionario_id):
        flash_erro("Funcionário não encontrado para esta empresa.")
        return redirect("/admin" if usuario_eh_admin(admin_id) else "/dashboard")

    data_inicio_raw = request.args.get("data_inicio")
    data_fim_raw = request.args.get("data_fim")
    data_inicio, data_fim, erro_periodo = validar_periodo(data_inicio_raw, data_fim_raw)
    destino = f"/admin/funcionarios/{funcionario_id}"
    if erro_periodo:
        flash_erro(erro_periodo)
        return redirect(destino)

    funcionario = obter_usuario(funcionario_id)
    if not funcionario:
        flash_erro("Funcionário não encontrado.")
        return redirect("/admin")

    esperado_min = funcionario.get("horas_diarias_esperadas_min") or 8 * 60
    registros_db = buscar_registros_usuario(funcionario_id, data_inicio, data_fim)
    if not registros_db:
        flash_aviso("Não há registros no período selecionado para exportar.")
        return redirect(destino)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"
    ws.append([periodo_para_texto(data_inicio, data_fim)])
    ws.append([f"Funcionário: {nome_usuario_para_exibicao(funcionario)}"])
    ws.append([f"Empresa: {empresa_normalizada(funcionario) or '-'}"])
    ws.append([])
    headers = ["Data", "Entrada", "Saída Almoço", "Volta Almoço", "Saída Final", "Total Trabalhado", "Saldo do Dia"]
    ws.append(headers)

    for cell in ws[5]:
        cell.font = Font(bold=True)

    for registro in registros_db:
        total_linha = calcular_total_registro(registro)
        em_aberto = not bool(parse_hora(registro[6]))
        saldo_delta = None if em_aberto else calcular_saldo_dia(total_linha, esperado_min)
        saldo_str = "em aberto" if em_aberto else (formatar_horas_minutos(saldo_delta) if saldo_delta is not None else "-")
        ws.append([registro[2], registro[3] or "", registro[4] or "", registro[5] or "", registro[6] or "", formatar_duracao_sem_sinal(total_linha), saldo_str])

    ws.freeze_panes = "A6"
    for col_idx, _ in enumerate(headers, start=1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            value = row[0]
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"registros_ponto_funcionario_{funcionario_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/funcionarios/<int:funcionario_id>/export/pdf")
def admin_export_pdf(funcionario_id: int):
    if not usuario_logado():
        return redirect("/")
    admin_id = session["user_id"]
    if not admin_pode_acessar_funcionario(admin_id, funcionario_id):
        flash_erro("Funcionário não encontrado para esta empresa.")
        return redirect("/admin" if usuario_eh_admin(admin_id) else "/dashboard")

    data_inicio_raw = request.args.get("data_inicio")
    data_fim_raw = request.args.get("data_fim")
    data_inicio, data_fim, erro_periodo = validar_periodo(data_inicio_raw, data_fim_raw)
    destino = f"/admin/funcionarios/{funcionario_id}"
    if erro_periodo:
        flash_erro(erro_periodo)
        return redirect(destino)

    funcionario = obter_usuario(funcionario_id)
    if not funcionario:
        flash_erro("Funcionário não encontrado.")
        return redirect("/admin")

    esperado_min = funcionario.get("horas_diarias_esperadas_min") or 8 * 60
    registros_db = buscar_registros_usuario(funcionario_id, data_inicio, data_fim)
    if not registros_db:
        flash_aviso("Não há registros no período selecionado para exportar.")
        return redirect(destino)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm, leftMargin=16 * mm, rightMargin=16 * mm, pageCompression=0, title="Relatório de Ponto")
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleCenter", parent=styles["Title"], alignment=TA_CENTER))
    story = [
        Paragraph("Relatório de Ponto", styles["TitleCenter"]),
        Spacer(1, 4 * mm),
        Paragraph(f"<b>Funcionário:</b> {nome_usuario_para_exibicao(funcionario)}", styles["Normal"]),
        Paragraph(f"<b>Empresa:</b> {empresa_normalizada(funcionario) or '-'}", styles["Normal"]),
        Paragraph(f"<b>{periodo_para_texto(data_inicio, data_fim)}</b>", styles["Normal"]),
        Spacer(1, 5 * mm),
    ]
    data_table = [["Data", "Entrada", "Saída Almoço", "Volta Almoço", "Saída Final", "Total", "Saldo"]]
    for registro in registros_db:
        total_linha = calcular_total_registro(registro)
        em_aberto = not bool(parse_hora(registro[6]))
        saldo_delta = None if em_aberto else calcular_saldo_dia(total_linha, esperado_min)
        saldo_str = "em aberto" if em_aberto else (formatar_horas_minutos(saldo_delta) if saldo_delta is not None else "-")
        data_table.append([registro[2], registro[3] or "", registro[4] or "", registro[5] or "", registro[6] or "", formatar_duracao_sem_sinal(total_linha), saldo_str])

    table = Table(data_table, repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")), ("FONTSIZE", (0, 0), (-1, -1), 8)]))
    story.append(table)
    doc.build(story)

    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"registros_ponto_funcionario_{funcionario_id}.pdf", mimetype="application/pdf")


@app.route("/perfil", methods=["GET", "POST"])
def perfil():
    if not usuario_logado():
        return redirect("/")

    if request.method == "POST":
        dados_perfil = {
            "nome_funcionario": request.form.get("nome_funcionario"),
            "nome_exibicao": request.form.get("nome_exibicao"),
            "horas_diarias_esperadas": request.form.get("horas_diarias_esperadas"),
        }
        if "nome_empresa" in request.form:
            dados_perfil["nome_empresa"] = request.form.get("nome_empresa")

        ok, erro = atualizar_perfil_usuario(
            session["user_id"],
            dados_perfil,
        )
        if not ok:
            flash_erro(erro or "Não foi possível salvar o perfil.")
            return redirect("/perfil")

        flash_ok("Perfil salvo com sucesso.")
        return redirect("/perfil")

    perfil_atual = obter_perfil_usuario(session["user_id"])
    horas_min = perfil_atual.get("horas_diarias_esperadas_min")
    horas_str = ""
    if horas_min is not None:
        horas_str = f"{int(horas_min // 60)}:{int(horas_min % 60):02d}"

    return render_template(
        "perfil.html",
        perfil=perfil_atual,
        horas_diarias_esperadas=horas_str,
    )


@app.route("/perfil/alterar-senha", methods=["POST"])
def alterar_senha():
    if not usuario_logado():
        return redirect("/")

    senha_atual = request.form.get("senha_atual") or ""
    nova_senha = request.form.get("nova_senha") or ""
    confirmar_nova = request.form.get("confirmar_nova_senha") or ""

    if not nova_senha.strip():
        flash_erro("Nova senha não pode ser vazia.")
        return redirect("/perfil")

    if nova_senha != confirmar_nova:
        flash_erro("Nova senha e confirmação não conferem.")
        return redirect("/perfil")

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(sql("SELECT password FROM usuarios WHERE id=?"), (session["user_id"],))
    row = cursor.fetchone()
    if not row:
        conn.close()
        flash_erro("Usuário não encontrado.")
        return redirect("/perfil")

    senha_hash = row[0]
    if not check_password_hash(senha_hash, senha_atual):
        conn.close()
        flash_erro("Senha atual incorreta.")
        return redirect("/perfil")

    nova_hash = generate_password_hash(nova_senha)
    cursor.execute(sql("UPDATE usuarios SET password=? WHERE id=?"), (nova_hash, session["user_id"]))
    conn.commit()
    conn.close()

    flash_ok("Senha alterada com sucesso.")
    return redirect("/perfil")


@app.route("/registros/<int:registro_id>/excluir", methods=["POST"])
def excluir_registro(registro_id: int):
    if not usuario_logado():
        return redirect("/")

    if excluir_registro_usuario(session["user_id"], registro_id):
        flash_ok("Registro excluído com sucesso.")
    else:
        flash_erro("Registro não encontrado.")

    return redirect("/dashboard")


@app.route("/registros/<int:registro_id>/editar", methods=["POST"])
def editar_registro(registro_id: int):
    if not usuario_logado():
        return redirect("/")

    user_id = session["user_id"]
    if not usuario_pode_editar_registro_proprio(user_id, registro_id):
        flash_erro("Registro nÃ£o encontrado.")
        return redirect("/dashboard")

    dados, erro = dados_correcao_do_formulario()
    if erro:
        flash_erro(erro)
        return redirect("/dashboard")

    if atualizar_registro_manual(registro_id, dados or {}):
        flash_ok("Registro corrigido com sucesso.")
    else:
        flash_erro("Registro nÃ£o encontrado.")

    return redirect("/dashboard")


@app.route("/admin/registros/<int:registro_id>/editar", methods=["POST"])
def admin_editar_registro(registro_id: int):
    if not usuario_logado():
        return redirect("/")

    admin_id = session["user_id"]
    registro = obter_registro(registro_id)
    funcionario_id = int(registro["user_id"]) if registro else None
    destino = f"/admin/funcionarios/{funcionario_id}" if funcionario_id else "/admin"

    if not admin_pode_editar_registro(admin_id, registro_id):
        flash_erro("Registro nÃ£o encontrado para esta empresa.")
        return redirect("/admin" if usuario_eh_admin(admin_id) else "/dashboard")

    dados, erro = dados_correcao_do_formulario()
    if erro:
        flash_erro(erro)
        return redirect(destino)

    if atualizar_registro_manual(registro_id, dados or {}):
        flash_ok("Registro corrigido com sucesso.")
    else:
        flash_erro("Registro nÃ£o encontrado.")

    return redirect(destino)


@app.route("/perfil/excluir-conta", methods=["POST"])
def excluir_conta():
    if not usuario_logado():
        return redirect("/")

    confirmacao = (request.form.get("confirmacao") or "").strip()
    if confirmacao != "EXCLUIR":
        flash_erro("Confirmação inválida. Digite EXCLUIR para apagar sua conta.")
        return redirect("/perfil")

    user_id = session["user_id"]
    excluir_conta_usuario(user_id)

    session.clear()
    flash_ok("Conta excluída com sucesso.")
    return redirect("/")


@app.route("/bater")
def bater():
    """
    Registra próxima etapa do ponto:
    - Entrada
    - Saída almoço
    - Volta almoço
    - Saída final
    """
    if not usuario_logado():
        return redirect("/")

    registrar_ponto(session["user_id"], agora())
    flash_ok("Ponto registrado com sucesso.")

    return redirect("/dashboard")


# ==========================================================
# API REST (JSON)
# ==========================================================

@app.route("/api/profile", methods=["GET"])
def api_get_profile():
    erro = exigir_login_api()
    if erro:
        return erro

    perfil = obter_perfil_usuario(session["user_id"])
    return api_ok({"profile": perfil_para_json(perfil)})


@app.route("/api/profile", methods=["PUT"])
def api_put_profile():
    erro = exigir_login_api()
    if erro:
        return erro

    dados = request.get_json(silent=True) or {}
    ok, mensagem = atualizar_perfil_usuario(session["user_id"], dados)
    if not ok:
        return api_erro(mensagem or "Dados inválidos", 400)

    perfil = obter_perfil_usuario(session["user_id"])
    return api_ok({"profile": perfil_para_json(perfil)})


@app.route("/api/registros", methods=["GET"])
def api_listar_registros():
    erro = exigir_login_api()
    if erro:
        return erro

    perfil = obter_perfil_usuario(session["user_id"])
    esperado_min = perfil.get("horas_diarias_esperadas_min")
    if esperado_min is None:
        esperado_min = 8 * 60

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        sql("""
            SELECT id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final,
                   corrigido_manual, motivo_correcao, corrigido_em
            FROM registros
            WHERE user_id=?
            ORDER BY data DESC, id DESC
        """),
        (session["user_id"],),
    )
    registros_db = cursor.fetchall()
    conn.close()

    registros: list[Dict[str, Any]] = []
    for registro in registros_db:
        total_linha = calcular_total_registro(registro)
        saida_final = registro[6]
        em_aberto = not bool(parse_hora(saida_final))
        saldo_delta = None if em_aberto else calcular_saldo_dia(total_linha, esperado_min)
        saldo_str = "em aberto" if em_aberto else (formatar_horas_minutos(saldo_delta) if saldo_delta is not None else "-")

        registros.append(
            {
                "id": registro[0],
                "data": registro[2],
                "entrada_manha": registro[3],
                "saida_almoco": registro[4],
                "volta_almoco": registro[5],
                "saida_final": registro[6],
                "total_seconds": int(total_linha.total_seconds()),
                "saldo": saldo_str,
                "em_aberto": em_aberto,
                "corrigido_manual": bool(registro[7]) if len(registro) > 7 else False,
                "motivo_correcao": registro[8] if len(registro) > 8 else None,
                "corrigido_em": registro[9] if len(registro) > 9 else None,
            }
        )

    return api_ok({"registros": registros})


@app.route("/api/ponto", methods=["POST"])
def api_registrar_ponto():
    erro = exigir_login_api()
    if erro:
        return erro

    resultado = registrar_ponto(session["user_id"], agora())
    return api_ok({"ponto": resultado}, 201)


@app.route("/api/registros/<int:registro_id>", methods=["DELETE"])
def api_excluir_registro(registro_id: int):
    erro = exigir_login_api()
    if erro:
        return erro

    if not excluir_registro_usuario(session["user_id"], registro_id):
        return api_erro("Registro não encontrado", 404)
    return api_ok({"deleted_id": registro_id})


@app.route("/api/account", methods=["DELETE"])
def api_excluir_conta():
    erro = exigir_login_api()
    if erro:
        return erro

    dados = request.get_json(silent=True) or {}
    confirmacao = (dados.get("confirmacao") or "").strip()
    if confirmacao != "EXCLUIR":
        return api_erro("Confirmação inválida. Envie confirmacao=EXCLUIR.", 400)

    user_id = session["user_id"]
    excluir_conta_usuario(user_id)
    session.clear()

    return api_ok({"deleted": True})


@app.route("/logout")
def logout():
    """
    Encerra sessão do usuário.
    """
    session.clear()
    return redirect("/")


# ==========================================================
# PWA (manifest + service worker)
# ==========================================================

@app.route("/manifest.json")
def pwa_manifest():
    return send_from_directory(app.static_folder, "manifest.json", mimetype="application/manifest+json")


@app.route("/service-worker.js")
def pwa_service_worker():
    resp = send_from_directory(app.static_folder, "service-worker.js", mimetype="application/javascript")
    # Service worker should not be cached aggressively during development
    resp.headers["Cache-Control"] = "no-cache"
    return resp


# ==========================================================
# ERROS (PÁGINAS AMIGÁVEIS)
# ==========================================================

@app.errorhandler(404)
def page_not_found(e):
    destino = "/dashboard" if usuario_logado() else "/"
    return render_template("404.html", destino=destino), 404


@app.errorhandler(500)
def internal_error(e):
    destino = "/dashboard" if usuario_logado() else "/"
    return render_template("500.html", destino=destino), 500


@app.route("/export/excel")
def export_excel():
    """
    Exporta registros do usuário logado em Excel (.xlsx).
    """
    if not usuario_logado():
        return redirect("/")

    data_inicio_raw = request.args.get("data_inicio")
    data_fim_raw = request.args.get("data_fim")
    data_inicio, data_fim, erro_periodo = validar_periodo(data_inicio_raw, data_fim_raw)
    if erro_periodo:
        flash_erro(erro_periodo)
        qs = ""
        if data_inicio_raw or data_fim_raw:
            qs = f"?data_inicio={data_inicio_raw or ''}&data_fim={data_fim_raw or ''}"
        return redirect("/dashboard" + qs)

    user_id = session["user_id"]
    perfil = obter_perfil_usuario(user_id)
    esperado_min = perfil.get("horas_diarias_esperadas_min")
    if esperado_min is None:
        esperado_min = 8 * 60

    conn = conectar()
    cursor = conn.cursor()
    sql_query, params_extra = aplicar_filtro_periodo_sql(
        """
            SELECT id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final,
                   corrigido_manual, motivo_correcao, corrigido_em
            FROM registros
            WHERE user_id=?
        """,
        data_inicio,
        data_fim,
    )
    sql_query += "\n            ORDER BY data DESC, id DESC"
    cursor.execute(sql(sql_query), (user_id,) + params_extra)
    registros_db = cursor.fetchall()
    conn.close()

    if not registros_db:
        flash_aviso("Não há registros no período selecionado para exportar.")
        qs = ""
        if data_inicio or data_fim:
            qs = f"?data_inicio={data_inicio or ''}&data_fim={data_fim or ''}"
        return redirect("/dashboard" + qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    if data_inicio and data_fim:
        periodo = f"Período: {datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    elif data_inicio:
        periodo = f"Período: a partir de {datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    elif data_fim:
        periodo = f"Período: até {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    else:
        periodo = "Período: todos os registros"

    ws.append([periodo])
    ws.append([])

    headers = [
        "Data",
        "Entrada",
        "Saída Almoço",
        "Volta Almoço",
        "Saída Final",
        "Total Trabalhado",
        "Saldo do Dia",
    ]
    ws.append(headers)

    ws["A1"].font = Font(bold=True)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    for registro in registros_db:
        total_linha = calcular_total_registro(registro)
        saida_final = registro[6]
        em_aberto = not bool(parse_hora(saida_final))
        saldo_delta = None if em_aberto else calcular_saldo_dia(total_linha, esperado_min)
        saldo_str = "em aberto" if em_aberto else (formatar_horas_minutos(saldo_delta) if saldo_delta is not None else "-")

        ws.append(
            [
                registro[2],
                registro[3] or "",
                registro[4] or "",
                registro[5] or "",
                registro[6] or "",
                formatar_duracao_sem_sinal(total_linha),
                saldo_str,
            ]
        )

    ws.freeze_panes = "A4"

    # Ajuste simples de largura das colunas
    for col_idx, _ in enumerate(headers, start=1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            value = row[0]
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="registros_ponto.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/pdf")
def export_pdf():
    """
    Exporta registros do usuário logado em PDF.
    """
    if not usuario_logado():
        return redirect("/")

    data_inicio_raw = request.args.get("data_inicio")
    data_fim_raw = request.args.get("data_fim")
    data_inicio, data_fim, erro_periodo = validar_periodo(data_inicio_raw, data_fim_raw)
    if erro_periodo:
        flash_erro(erro_periodo)
        qs = ""
        if data_inicio_raw or data_fim_raw:
            qs = f"?data_inicio={data_inicio_raw or ''}&data_fim={data_fim_raw or ''}"
        return redirect("/dashboard" + qs)

    user_id = session["user_id"]
    perfil = obter_perfil_usuario(user_id)
    esperado_min = perfil.get("horas_diarias_esperadas_min")
    if esperado_min is None:
        esperado_min = 8 * 60

    conn = conectar()
    cursor = conn.cursor()
    sql_query, params_extra = aplicar_filtro_periodo_sql(
        """
            SELECT id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final,
                   corrigido_manual, motivo_correcao, corrigido_em
            FROM registros
            WHERE user_id=?
        """,
        data_inicio,
        data_fim,
    )
    sql_query += "\n            ORDER BY data DESC, id DESC"
    cursor.execute(sql(sql_query), (user_id,) + params_extra)
    registros_db = cursor.fetchall()
    conn.close()

    if not registros_db:
        flash_aviso("Não há registros no período selecionado para exportar.")
        qs = ""
        if data_inicio or data_fim:
            qs = f"?data_inicio={data_inicio or ''}&data_fim={data_fim or ''}"
        return redirect("/dashboard" + qs)

    if data_inicio and data_fim:
        periodo = f"{datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    elif data_inicio:
        periodo = f"a partir de {datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    elif data_fim:
        periodo = f"até {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    else:
        periodo = "todos os registros"

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        pageCompression=0,
        title="Relatório de Ponto",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleCenter", parent=styles["Title"], alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=9, leading=11))
    styles.add(ParagraphStyle(name="SmallRight", parent=styles["Small"], alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name="SmallLeft", parent=styles["Small"], alignment=TA_LEFT))
    story = []

    data_geracao = agora().strftime("%d/%m/%Y %H:%M")

    nome_funcionario = (perfil.get("nome_funcionario") or "").strip()
    nome_exibicao = (perfil.get("nome_exibicao") or "").strip()
    nome_empresa = (perfil.get("nome_empresa") or "").strip()

    nome_prioritario = nome_exibicao or nome_funcionario or "não informado"

    story.append(Paragraph("Relatório de Ponto", styles["TitleCenter"]))
    story.append(Spacer(1, 4 * mm))

    header_data = [
        [
            Paragraph(f"<b>Funcionário:</b> {nome_prioritario}", styles["SmallLeft"]),
            Paragraph(f"<b>Data de geração:</b> {data_geracao}", styles["SmallRight"]),
        ],
        [
            Paragraph(f"<b>Nome de exibição:</b> {nome_exibicao or '-'}", styles["SmallLeft"]),
            Paragraph(f"<b>Período:</b> {periodo}", styles["SmallRight"]),
        ],
    ]
    if nome_empresa:
        header_data.append(
            [
                Paragraph(f"<b>Empresa:</b> {nome_empresa}", styles["SmallLeft"]),
                Paragraph("", styles["SmallRight"]),
            ]
        )
    if nome_exibicao and nome_funcionario and nome_exibicao != nome_funcionario:
        header_data.append(
            [
                Paragraph(f"<b>Nome do funcionário:</b> {nome_funcionario}", styles["SmallLeft"]),
                Paragraph("", styles["SmallRight"]),
            ]
        )

    header_table = Table(header_data, colWidths=[(A4[0] - doc.leftMargin - doc.rightMargin) * 0.55, (A4[0] - doc.leftMargin - doc.rightMargin) * 0.45])
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 5 * mm))

    data_table = [
        ["Data", "Entrada", "Saída Almoço", "Volta Almoço", "Saída Final", "Total Trabalhado", "Saldo do Dia"]
    ]

    total_periodo = timedelta()
    banco_periodo = timedelta()

    for registro in registros_db:
        total_linha = calcular_total_registro(registro)
        saida_final = registro[6]
        em_aberto = not bool(parse_hora(saida_final))
        saldo_delta = None if em_aberto else calcular_saldo_dia(total_linha, esperado_min)
        saldo_str = "em aberto" if em_aberto else (formatar_horas_minutos(saldo_delta) if saldo_delta is not None else "-")

        if not em_aberto:
            total_periodo += total_linha
            if saldo_delta is not None:
                banco_periodo += saldo_delta

        data_table.append(
            [
                registro[2],
                registro[3] or "",
                registro[4] or "",
                registro[5] or "",
                registro[6] or "",
                formatar_duracao_sem_sinal(total_linha),
                saldo_str,
            ]
        )

    usable_width = A4[0] - doc.leftMargin - doc.rightMargin
    col_widths = [
        usable_width * 0.12,  # data
        usable_width * 0.11,  # entrada
        usable_width * 0.13,  # saida almoco
        usable_width * 0.13,  # volta almoco
        usable_width * 0.11,  # saida final
        usable_width * 0.14,  # total
        usable_width * 0.16,  # saldo
    ]

    table = Table(data_table, repeatRows=1, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f1f5f9")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (1, 0), (-1, 0), "CENTER"),
                ("ALIGN", (1, 1), (-2, -1), "CENTER"),
                ("ALIGN", (-2, 1), (-1, -1), "RIGHT"),
            ]
        )
    )

    story.append(table)
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph(f"<b>Total de horas trabalhadas:</b> {formatar_duracao_sem_sinal(total_periodo)}", styles["Small"]))
    story.append(Paragraph(f"<b>Banco de horas do período:</b> {formatar_banco_horas(banco_periodo)}", styles["Small"]))
    doc.build(story)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="registros_ponto.pdf",
        mimetype="application/pdf",
    )


# ==========================================================
# INICIALIZAÇÃO
# ==========================================================

if __name__ == "__main__":
    criar_banco()
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
