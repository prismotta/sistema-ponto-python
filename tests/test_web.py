import pytest
import os
from web.app import app, criar_banco, parse_hora
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
import json

@pytest.fixture
def client():
    # Usa banco de teste separado
    app.config["TESTING"] = True
    app.config["SECRET_KEY"] = "test"

    # Remove banco antigo se existir
    if os.path.exists("web/database.db"):
        os.remove("web/database.db")

    criar_banco()

    with app.test_client() as client:
        yield client


def test_login_valido(client):
    # Criar usuário
    client.post("/register", data={
        "username": "teste",
        "password": "1234"
    })

    # Fazer login
    response = client.post("/", data={
        "username": "teste",
        "password": "1234"
    }, follow_redirects=False)

    # Verifica redirecionamento para dashboard
    assert response.status_code == 302
    assert "/dashboard" in response.headers["Location"]

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM usuarios WHERE username=?", ("teste",))
    role = cursor.fetchone()[0]
    conn.close()
    assert role == "user"

def test_login_invalido(client):
    # Criar usuário válido
    client.post("/register", data={
        "username": "teste",
        "password": "1234"
    })

    # Tentar login com senha errada
    response = client.post("/", data={
        "username": "teste",
        "password": "senha_errada"
    }, follow_redirects=True)

    # Deve permanecer na página de login
    assert response.status_code == 200

    # Verifica mensagem de erro na resposta
    assert b"Usu\xc3\xa1rio ou senha inv\xc3\xa1lidos" in response.data

def test_dashboard_protegido_sem_login(client):
    response = client.get("/dashboard", follow_redirects=False)

    # Deve redirecionar
    assert response.status_code == 302

    # Deve redirecionar para login
    assert "/" in response.headers["Location"]

def test_logout(client):
    # Criar usuário
    client.post("/register", data={
        "username": "teste",
        "password": "1234"
    })

    # Fazer login
    client.post("/", data={
        "username": "teste",
        "password": "1234"
    })

    # Garantir que dashboard funciona logado
    response_dashboard = client.get("/dashboard")
    assert response_dashboard.status_code == 200

    # Fazer logout
    client.get("/logout")

    # Tentar acessar dashboard novamente
    response = client.get("/dashboard", follow_redirects=False)

    # Deve redirecionar para login
    assert response.status_code == 302
    assert "/" in response.headers["Location"]

def test_fluxo_bater_ponto(client):
    # Criar usuário
    client.post("/register", data={
        "username": "teste",
        "password": "1234"
    })

    # Login
    client.post("/", data={
        "username": "teste",
        "password": "1234"
    })

    # 1ª batida → entrada_manha
    client.get("/bater")

    # 2ª batida → saida_almoco
    client.get("/bater")

    # 3ª batida → volta_almoco
    client.get("/bater")

    # 4ª batida → saida_final
    client.get("/bater")

    # Acessar dashboard
    response = client.get("/dashboard")

    assert response.status_code == 200

    # Verificar que as colunas aparecem preenchidas
    assert b"Sa\xc3\xadda Final" in response.data or b"Sa\xc3\xadda Final" in response.data


def test_dashboard_mostra_historico_de_dias_anteriores(client):
    # Criar usuário e logar
    client.post("/register", data={
        "username": "teste",
        "password": "1234"
    })
    client.post("/", data={
        "username": "teste",
        "password": "1234"
    })

    # Cria ao menos um registro de hoje
    client.get("/bater")

    # Insere manualmente um registro completo de ontem para simular histórico
    import sqlite3
    from datetime import datetime, timedelta

    ontem = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        # Formatos legados (sem segundos) não devem quebrar o histórico.
        (1, ontem, "08:00", "12:00", "13:00", "17:00")
    )
    conn.commit()
    conn.close()

    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b"Saldo do Dia" in response.data
    assert ontem.encode("utf-8") in response.data


def test_parse_hora_aceita_formatos_legados():
    assert parse_hora(None) is None
    assert parse_hora("") is None

    assert parse_hora("08:30").strftime("%H:%M:%S") == "08:30:00"
    assert parse_hora("08:30:15").strftime("%H:%M:%S") == "08:30:15"
    assert parse_hora("08:30:15.123456").microsecond == 123456


def test_criar_banco_migra_registros_legado_sem_colunas_novas():
    app.config["TESTING"] = True
    app.config["SECRET_KEY"] = "test"

    if os.path.exists("web/database.db"):
        os.remove("web/database.db")

    import sqlite3

    # Simula um banco legado: tabela `registros` existe mas não tem as colunas novas
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)
    cursor.execute("""
        CREATE TABLE registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            data TEXT,
            entrada_manha TEXT,
            FOREIGN KEY(user_id) REFERENCES usuarios(id)
        )
    """)
    conn.commit()
    conn.close()

    criar_banco()

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(registros)")
    colunas = {row[1] for row in cursor.fetchall()}
    conn.close()

    assert {"entrada_manha", "saida_almoco", "volta_almoco", "saida_final"}.issubset(colunas)

    with app.test_client() as client:
        client.post("/register", data={"username": "teste", "password": "1234"})
        client.post("/", data={"username": "teste", "password": "1234"})

        # Deve conseguir fazer batidas sem erro, mesmo partindo de schema legado
        client.get("/bater")
        client.get("/bater")

        response = client.get("/dashboard")
        assert response.status_code == 200


def _criar_usuario_e_logar(client, username: str = "teste", password: str = "1234"):
    client.post("/register", data={"username": username, "password": password})
    client.post("/", data={"username": username, "password": password})


def _definir_role(user_id: int, role: str):
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("UPDATE usuarios SET role=? WHERE id=?", (role, user_id))
    conn.commit()
    conn.close()


def _definir_empresa(user_id: int, empresa: str, nome: str):
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE usuarios SET nome_empresa=?, nome_funcionario=?, horas_diarias_esperadas_min=? WHERE id=?",
        (empresa, nome, 8 * 60, user_id),
    )
    conn.commit()
    conn.close()


def _inserir_registro(user_id: int, data: str, entrada="08:00", saida_almoco="12:00", volta_almoco="13:00", saida_final="17:00"):
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (user_id, data, entrada, saida_almoco, volta_almoco, saida_final),
    )
    registro_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return registro_id


def _dados_correcao(**overrides):
    dados = {
        "data": "2026-04-10",
        "entrada_manha": "08:00",
        "saida_almoco": "12:00",
        "volta_almoco": "13:00",
        "saida_final": "18:00",
        "motivo_correcao": "Esqueci de bater saída do almoço",
    }
    dados.update(overrides)
    return dados


def _buscar_registro_db(registro_id: int):
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT id, user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final,
               corrigido_manual, motivo_correcao, corrigido_em
        FROM registros
        WHERE id=?
        """,
        (registro_id,),
    )
    row = cursor.fetchone()
    conn.close()
    return row


def _buscar_convite_por_email(email: str):
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id, email, empresa, token, usado FROM invites WHERE email=?", (email,))
    row = cursor.fetchone()
    conn.close()
    return row


def test_salvar_perfil(client):
    _criar_usuario_e_logar(client)

    response = client.post(
        "/perfil",
        data={
            "nome_funcionario": "João da Silva",
            "nome_exibicao": "João",
            "nome_empresa": "Empresa X",
            "horas_diarias_esperadas": "08:30",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT nome_funcionario, nome_exibicao, nome_empresa, horas_diarias_esperadas_min FROM usuarios WHERE id=1")
    row = cursor.fetchone()
    conn.close()

    assert row == ("João da Silva", "João", "Empresa X", 8 * 60 + 30)


def test_perfil_exige_empresa_quando_ainda_nao_definida(client):
    _criar_usuario_e_logar(client)

    response = client.post(
        "/perfil",
        data={"nome_funcionario": "João", "nome_empresa": "", "horas_diarias_esperadas": "8"},
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Informe a empresa para continuar.".encode("utf-8") in response.data


def test_perfil_nao_permite_alterar_empresa_ja_definida(client):
    _criar_usuario_e_logar(client)
    client.post(
        "/perfil",
        data={"nome_funcionario": "João", "nome_empresa": "Empresa Original", "horas_diarias_esperadas": "8"},
        follow_redirects=True,
    )

    response = client.post(
        "/perfil",
        data={"nome_funcionario": "João", "nome_empresa": "Empresa Alterada", "horas_diarias_esperadas": "8"},
        follow_redirects=True,
    )

    assert response.status_code == 200
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT nome_empresa FROM usuarios WHERE id=1")
    empresa = cursor.fetchone()[0]
    conn.close()
    assert empresa == "Empresa Original"
    assert b"Empresa Alterada" not in response.data


def test_deletar_registro_so_afeta_usuario_logado(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    client.get("/bater")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM registros WHERE user_id=1")
    registro_id = cursor.fetchone()[0]
    conn.close()

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")

    # Tentativa de excluir registro de outro usuário não deve apagar nada
    client.post(f"/registros/{registro_id}/excluir", follow_redirects=True)

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM registros WHERE id=?", (registro_id,))
    count = cursor.fetchone()[0]
    conn.close()
    assert count == 1


def test_deletar_registro_do_proprio_usuario(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    client.get("/bater")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM registros WHERE user_id=1")
    registro_id = cursor.fetchone()[0]
    conn.close()

    client.post(f"/registros/{registro_id}/excluir", follow_redirects=True)

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM registros WHERE id=?", (registro_id,))
    count = cursor.fetchone()[0]
    conn.close()
    assert count == 0


def test_calcular_saldo_extra_faltante_e_em_aberto(client):
    _criar_usuario_e_logar(client)

    # Define esperado: 8h
    client.post("/perfil", data={"horas_diarias_esperadas": "8"}, follow_redirects=True)

    hoje = datetime.now().strftime("%Y-%m-%d")
    ontem = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    # Ontem: 9h trabalhadas -> +1h
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, ontem, "08:00", "12:00", "13:00", "18:00"),
    )
    # Dois dias atrás: 7h trabalhadas -> -1h
    anteontem = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d")
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, anteontem, "08:00", "12:00", "13:00", "16:00"),
    )
    # Hoje: sem saída final -> em aberto
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha)
        VALUES (?, ?, ?)
        """,
        (1, hoje, "08:00"),
    )
    conn.commit()
    conn.close()

    response = client.get("/dashboard")
    assert response.status_code == 200
    assert b"+1h 00m" in response.data
    assert b"-1h 00m" in response.data
    assert b"em aberto" in response.data


def test_dashboard_exibe_nome_funcionario_ou_nao_informado(client):
    _criar_usuario_e_logar(client)

    # Sem nome no perfil
    response = client.get("/dashboard")
    assert response.status_code == 200
    assert "Funcionário: não informado".encode("utf-8") in response.data

    # Com nome no perfil
    client.post("/perfil", data={"nome_funcionario": "Maria", "horas_diarias_esperadas": "8"}, follow_redirects=True)
    response = client.get("/dashboard")
    assert response.status_code == 200
    assert "Funcionário: Maria".encode("utf-8") in response.data


def test_dashboard_prioriza_nome_exibicao(client):
    _criar_usuario_e_logar(client)
    client.post(
        "/perfil",
        data={"nome_funcionario": "Priscila F. Motta", "nome_exibicao": "Pri", "horas_diarias_esperadas": "8"},
        follow_redirects=True,
    )

    response = client.get("/dashboard")
    assert response.status_code == 200
    assert "Funcionário: Pri".encode("utf-8") in response.data


def test_dashboard_usa_primeiro_nome_quando_nome_exibicao_vazio(client):
    _criar_usuario_e_logar(client)
    client.post(
        "/perfil",
        data={"nome_funcionario": "Priscila F. Motta", "nome_exibicao": "", "horas_diarias_esperadas": "8"},
        follow_redirects=True,
    )

    response = client.get("/dashboard")
    assert response.status_code == 200
    assert "Funcionário: Priscila".encode("utf-8") in response.data


def test_dashboard_botao_registrar_ponto(client):
    _criar_usuario_e_logar(client)
    response = client.get("/dashboard")
    assert response.status_code == 200
    html = response.data
    assert "Registrar entrada".encode("utf-8") in response.data
    assert b'href="/bater"' in html
    assert b"btn-ponto-principal" in html
    assert b"btn-ponto-principal-texto" in html
    assert b"btn-ponto-principal-spinner" in html
    assert b"data-loading-text=\"Registrando...\"" in html
    assert b"addEventListener('click', function(event)" in html
    assert b"if (navigator.vibrate)" in html
    assert b"navigator.vibrate(50)" in html
    assert b"classList.add('disabled')" in html
    assert b"aria-disabled" in html
    assert b"event.preventDefault()" in html
    assert html.index(b"navigator.vibrate(50)") < html.index(b"ativarLoadingPonto();")
    assert html.index(b"navigator.vibrate(50)") < html.index(b"event.preventDefault()")
    assert b"btn btn-outline-success" in html
    assert b"btn btn-outline-danger" in html
    assert b'id="btnEsteMes"' in html


def test_usuario_comum_nao_acessa_admin(client):
    _criar_usuario_e_logar(client, "funcionario", "1234")
    _definir_empresa(1, "Empresa A", "Funcionario A")

    response = client.get("/admin", follow_redirects=False)

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/dashboard")


def test_admin_acessa_admin_e_ve_funcionarios_da_propria_empresa(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_a", "1234")
    _definir_empresa(2, "Empresa A", "Funcionario A")
    _inserir_registro(2, "2026-04-10")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})
    response = client.get("/admin")

    assert response.status_code == 200
    assert "Funcionarios".encode("utf-8") in response.data or "Funcionários".encode("utf-8") in response.data
    assert b"Funcionario A" in response.data
    assert b"Empresa A" in response.data
    assert b">1<" in response.data


def test_admin_renderiza_acentuacao_em_utf8(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_a", "1234")
    _definir_empresa(2, "Empresa A", "Funcionario A")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})
    response = client.get("/admin")

    assert response.status_code == 200
    assert b'<meta charset="UTF-8">' in response.data
    for palavra in ("Ações", "Usuário", "Histórico", "Administração"):
        assert palavra.encode("utf-8") in response.data


def test_admin_sem_empresa_nao_ve_lista(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_sem_empresa", "1234")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})
    response = client.get("/admin")

    assert response.status_code == 200
    assert "Empresa não definida".encode("utf-8") in response.data
    assert "Defina sua empresa no Perfil para usar o painel admin.".encode("utf-8") in response.data
    assert b"funcionario_sem_empresa" not in response.data


def test_setup_admin_sem_login_nao_acessa(client, monkeypatch):
    monkeypatch.setenv("ALLOW_ADMIN_SETUP", "true")

    response = client.get("/setup-admin", follow_redirects=False)

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/")


def test_setup_admin_sem_variavel_bloqueia(client, monkeypatch):
    monkeypatch.delenv("ALLOW_ADMIN_SETUP", raising=False)
    _criar_usuario_e_logar(client, "funcionario", "1234")
    _definir_empresa(1, "Empresa A", "Funcionario A")

    response = client.get("/setup-admin", follow_redirects=False)

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/dashboard")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM usuarios WHERE id=1")
    role = cursor.fetchone()[0]
    conn.close()
    assert role == "user"


def test_setup_admin_com_variavel_e_login_promove_usuario_atual(client, monkeypatch):
    monkeypatch.setenv("ALLOW_ADMIN_SETUP", "true")
    _criar_usuario_e_logar(client, "funcionario", "1234")
    _definir_empresa(1, "Empresa A", "Funcionario A")

    promovido = client.get("/setup-admin", follow_redirects=False)
    assert promovido.status_code == 302
    assert promovido.headers["Location"].endswith("/admin")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, role FROM usuarios WHERE id=1")
    usuario = cursor.fetchone()
    conn.close()
    assert usuario == (1, "funcionario", "admin")

    response_admin = client.get("/admin")
    assert response_admin.status_code == 200

    response_dashboard = client.get("/dashboard")
    assert b'href="/admin"' in response_dashboard.data


def test_admin_nao_ve_funcionarios_de_outra_empresa(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_a", "1234")
    _definir_empresa(2, "Empresa A", "Funcionario A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_b", "1234")
    _definir_empresa(3, "Empresa B", "Funcionario B")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})
    response = client.get("/admin")

    assert response.status_code == 200
    assert b"Funcionario A" in response.data
    assert b"Funcionario B" not in response.data

    bloqueado = client.get("/admin/funcionarios/3", follow_redirects=False)
    assert bloqueado.status_code == 302
    assert bloqueado.headers["Location"].endswith("/admin")


def test_admin_nao_exporta_dados_de_outra_empresa(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_b", "1234")
    _definir_empresa(2, "Empresa B", "Funcionario B")
    _inserir_registro(2, "2026-04-10")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})

    excel = client.get("/admin/funcionarios/2/export/excel", follow_redirects=False)
    assert excel.status_code == 302
    assert excel.headers["Location"].endswith("/admin")

    pdf = client.get("/admin/funcionarios/2/export/pdf", follow_redirects=False)
    assert pdf.status_code == 302
    assert pdf.headers["Location"].endswith("/admin")


def test_admin_promove_usuario_da_mesma_empresa(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_a", "1234")
    _definir_empresa(2, "Empresa A", "Funcionario A")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})
    response = client.post("/admin/usuarios/2/promover-admin", follow_redirects=True)

    assert response.status_code == 200
    assert b"promovido a admin com sucesso" in response.data
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM usuarios WHERE id=2")
    role = cursor.fetchone()[0]
    conn.close()
    assert role == "admin"
    assert b"Tornar admin" not in response.data


def test_admin_nao_promove_usuario_de_outra_empresa(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_b", "1234")
    _definir_empresa(2, "Empresa B", "Funcionario B")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})
    response = client.post("/admin/usuarios/2/promover-admin", follow_redirects=False)

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/admin")
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM usuarios WHERE id=2")
    role = cursor.fetchone()[0]
    conn.close()
    assert role == "user"


def test_usuario_edita_proprio_registro_com_motivo_e_recalcula_saldo(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    _definir_empresa(1, "Empresa A", "Usuario 1")
    registro_id = _inserir_registro(1, "2026-04-10", saida_final="17:00")

    response = client.post(
        f"/registros/{registro_id}/editar",
        data=_dados_correcao(saida_final="18:00"),
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Registro corrigido com sucesso.".encode("utf-8") in response.data
    assert b"18:00" in response.data
    assert b"+1h 00m" in response.data
    assert b"Corrigido" in response.data
    assert "Esqueci de bater saída do almoço".encode("utf-8") in response.data

    registro = _buscar_registro_db(registro_id)
    assert registro[6] == "18:00"
    assert registro[7] == 1
    assert registro[8] == "Esqueci de bater saída do almoço"
    assert registro[9]


def test_usuario_nao_edita_registro_de_outro_usuario(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    _definir_empresa(1, "Empresa A", "Usuario 1")
    registro_id = _inserir_registro(1, "2026-04-10", saida_final="17:00")

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")
    _definir_empresa(2, "Empresa A", "Usuario 2")

    response = client.post(
        f"/registros/{registro_id}/editar",
        data=_dados_correcao(saida_final="18:00"),
        follow_redirects=False,
    )

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/dashboard")
    registro = _buscar_registro_db(registro_id)
    assert registro[6] == "17:00"
    assert registro[7] == 0


def test_admin_edita_registro_da_mesma_empresa(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario", "1234")
    _definir_empresa(2, "Empresa A", "Funcionario A")
    registro_id = _inserir_registro(2, "2026-04-10", saida_final="17:00")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})
    response = client.post(
        f"/admin/registros/{registro_id}/editar",
        data=_dados_correcao(saida_final="18:00", motivo_correcao="Ajuste feito pelo admin"),
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Registro corrigido com sucesso.".encode("utf-8") in response.data
    assert b"18:00" in response.data
    assert b"+1h 00m" in response.data
    registro = _buscar_registro_db(registro_id)
    assert registro[6] == "18:00"
    assert registro[7] == 1
    assert registro[8] == "Ajuste feito pelo admin"


def test_admin_nao_edita_registro_de_outra_empresa(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario", "1234")
    _definir_empresa(2, "Empresa B", "Funcionario B")
    registro_id = _inserir_registro(2, "2026-04-10", saida_final="17:00")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})
    response = client.post(
        f"/admin/registros/{registro_id}/editar",
        data=_dados_correcao(saida_final="18:00"),
        follow_redirects=False,
    )

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/admin")
    registro = _buscar_registro_db(registro_id)
    assert registro[6] == "17:00"
    assert registro[7] == 0


def test_motivo_correcao_e_obrigatorio(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    _definir_empresa(1, "Empresa A", "Usuario 1")
    registro_id = _inserir_registro(1, "2026-04-10", saida_final="17:00")

    response = client.post(
        f"/registros/{registro_id}/editar",
        data=_dados_correcao(saida_final="18:00", motivo_correcao=""),
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert "Informe o motivo da correção.".encode("utf-8") in response.data
    registro = _buscar_registro_db(registro_id)
    assert registro[6] == "17:00"
    assert registro[7] == 0


def test_usuario_comum_nao_ve_botao_e_nao_promove(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    _definir_empresa(1, "Empresa A", "Usuario 1")

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")
    _definir_empresa(2, "Empresa A", "Usuario 2")

    dashboard = client.get("/dashboard")
    assert b"Tornar admin" not in dashboard.data

    response = client.post("/admin/usuarios/1/promover-admin", follow_redirects=False)
    assert response.status_code == 302
    assert response.headers["Location"].endswith("/dashboard")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM usuarios WHERE id=1")
    role = cursor.fetchone()[0]
    conn.close()
    assert role == "user"


def test_convite_valido_cria_usuario_na_empresa_correta(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    response_convite = client.post("/admin/convites", data={"email": "novo@empresa.com"}, follow_redirects=True)
    assert response_convite.status_code == 200
    assert b"/register?token=" in response_convite.data

    convite = _buscar_convite_por_email("novo@empresa.com")
    assert convite is not None
    token = convite[3]

    client.get("/logout")
    response_get = client.get(f"/register?token={token}")
    assert response_get.status_code == 200
    assert b"Empresa A" in response_get.data

    response_register = client.post(
        "/register",
        data={"username": "novo", "password": "1234", "token": token},
        follow_redirects=False,
    )
    assert response_register.status_code == 302
    assert response_register.headers["Location"].endswith("/")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT username, nome_empresa, role FROM usuarios WHERE username=?", ("novo",))
    usuario = cursor.fetchone()
    cursor.execute("SELECT usado FROM invites WHERE token=?", (token,))
    usado = cursor.fetchone()[0]
    conn.close()

    assert usuario == ("novo", "Empresa A", "user")
    assert usado == 1


def test_convite_nao_pode_ser_reutilizado(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")
    client.post("/admin/convites", data={"email": "novo@empresa.com"})
    token = _buscar_convite_por_email("novo@empresa.com")[3]

    client.get("/logout")
    client.post("/register", data={"username": "novo", "password": "1234", "token": token})

    response = client.get(f"/register?token={token}")

    assert response.status_code == 200
    assert "Convite inválido ou já utilizado.".encode("utf-8") in response.data
    assert b'name="username"' not in response.data


def test_usuario_nao_consegue_mudar_empresa_do_convite_manualmente(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")
    client.post("/admin/convites", data={"email": "novo@empresa.com"})
    token = _buscar_convite_por_email("novo@empresa.com")[3]

    client.get("/logout")
    client.post(
        "/register",
        data={"username": "novo", "password": "1234", "token": token, "nome_empresa": "Empresa B"},
    )

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT nome_empresa FROM usuarios WHERE username=?", ("novo",))
    empresa = cursor.fetchone()[0]
    conn.close()
    assert empresa == "Empresa A"


def test_admin_cancela_convite_pendente(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")
    client.post("/admin/convites", data={"email": "novo@empresa.com"})
    convite_id = _buscar_convite_por_email("novo@empresa.com")[0]

    response = client.post(f"/admin/convites/{convite_id}/cancelar", follow_redirects=True)

    assert response.status_code == 200
    assert b"Convite cancelado com sucesso." in response.data
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT usado FROM invites WHERE id=?", (convite_id,))
    usado = cursor.fetchone()[0]
    conn.close()
    assert usado == 1
    assert b"novo@empresa.com" not in response.data


def test_exportacao_admin_respeita_funcionario_selecionado(client):
    _criar_usuario_e_logar(client, "admin", "1234")
    _definir_role(1, "admin")
    _definir_empresa(1, "Empresa A", "Admin A")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_a", "1234")
    _definir_empresa(2, "Empresa A", "Funcionario A")
    _inserir_registro(2, "2026-04-10", entrada="08:00", saida_almoco="12:00", volta_almoco="13:00", saida_final="17:00")

    client.get("/logout")
    _criar_usuario_e_logar(client, "funcionario_b", "1234")
    _definir_empresa(3, "Empresa A", "Funcionario B")
    _inserir_registro(3, "2026-04-11", entrada="09:00", saida_almoco="12:00", volta_almoco="13:00", saida_final="18:00")

    client.get("/logout")
    client.post("/", data={"username": "admin", "password": "1234"})

    response = client.get("/admin/funcionarios/2/export/excel")

    assert response.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(response.data))
    ws = wb.active
    valores = [cell for row in ws.iter_rows(values_only=True) for cell in row if cell is not None]
    assert any("Funcionario A" in str(valor) for valor in valores)
    assert "2026-04-10" in valores
    assert "2026-04-11" not in valores
    assert "09:00" not in valores


def test_api_sem_login_deve_falhar(client):
    resp = client.get("/api/profile")
    assert resp.status_code == 401
    assert resp.is_json
    assert resp.json["success"] is False


def test_pwa_manifest_e_service_worker_sao_servidos(client):
    resp_manifest = client.get("/manifest.json")
    assert resp_manifest.status_code == 200
    assert resp_manifest.is_json
    data = resp_manifest.get_json()
    assert data["name"] == "Sistema de Ponto"
    assert data["short_name"] == "Ponto"
    assert data["start_url"] == "/"

    resp_sw = client.get("/service-worker.js")
    assert resp_sw.status_code == 200
    assert b"service worker" in resp_sw.data.lower() or b"addEventListener" in resp_sw.data


def test_pwa_update_manual_envia_skip_waiting_e_recarrega_uma_vez(client):
    resp_js = client.get("/static/pwa-update.js")
    assert resp_js.status_code == 200
    js = resp_js.data

    assert b"let newWorker = null" in js
    assert b"registration.waiting" in js
    assert b"updatefound" in js
    assert b'state === "installed"' in js
    assert b'navigator.serviceWorker.controller' in js
    assert b'newWorker.postMessage({ type: "SKIP_WAITING" })' in js
    assert b"controllerchange" in js
    assert b"let refreshing = false" in js
    assert b"window.location.reload()" in js

    resp_sw = client.get("/service-worker.js")
    sw = resp_sw.data
    assert b"type === \"SKIP_WAITING\"" in sw
    assert b"self.skipWaiting()" in sw
    assert b".then(() => self.skipWaiting())" not in sw


def test_api_get_profile_logado(client):
    _criar_usuario_e_logar(client)
    client.put(
        "/api/profile",
        json={"nome_funcionario": "João da Silva", "nome_exibicao": "João", "nome_empresa": "Empresa X", "horas_diarias_esperadas": "8"},
    )
    resp = client.get("/api/profile")
    assert resp.status_code == 200
    assert resp.is_json
    assert resp.json["success"] is True
    assert resp.json["profile"]["nome_funcionario"] == "João da Silva"
    assert resp.json["profile"]["nome_exibicao"] == "João"


def test_api_put_profile_atualiza_perfil(client):
    _criar_usuario_e_logar(client)
    resp = client.put(
        "/api/profile",
        json={"nome_funcionario": "Maria Souza", "nome_exibicao": "Mari", "nome_empresa": "ACME", "horas_diarias_esperadas": "08:30"},
    )
    assert resp.status_code == 200
    assert resp.is_json
    assert resp.json["success"] is True
    assert resp.json["profile"]["horas_diarias_esperadas_min"] == 8 * 60 + 30


def test_api_listar_registros_apenas_do_usuario(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    client.post("/api/ponto")

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")
    client.post("/api/ponto")

    resp = client.get("/api/registros")
    assert resp.status_code == 200
    assert resp.is_json
    assert resp.json["success"] is True
    assert len(resp.json["registros"]) == 1


def test_usuario_comum_dashboard_mostra_apenas_proprios_dados(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    _definir_empresa(1, "Empresa A", "Usuario 1")
    _inserir_registro(1, "2026-04-10")

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")
    _definir_empresa(2, "Empresa A", "Usuario 2")
    _inserir_registro(2, "2026-04-11")

    response = client.get("/dashboard")

    assert response.status_code == 200
    assert b"2026-04-11" in response.data
    assert b"2026-04-10" not in response.data


def test_api_registrar_ponto(client):
    _criar_usuario_e_logar(client)
    resp = client.post("/api/ponto")
    assert resp.status_code == 201
    assert resp.is_json
    assert resp.json["success"] is True

    resp_list = client.get("/api/registros")
    assert resp_list.status_code == 200
    assert len(resp_list.json["registros"]) >= 1


def test_api_deletar_registro_proprio(client):
    _criar_usuario_e_logar(client)
    client.post("/api/ponto")

    resp_list = client.get("/api/registros")
    registro_id = resp_list.json["registros"][0]["id"]

    resp_del = client.delete(f"/api/registros/{registro_id}")
    assert resp_del.status_code == 200
    assert resp_del.is_json
    assert resp_del.json["success"] is True


def test_api_impede_deletar_registro_de_outro_usuario(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    client.post("/api/ponto")
    resp_list = client.get("/api/registros")
    registro_id = resp_list.json["registros"][0]["id"]

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")
    resp_del = client.delete(f"/api/registros/{registro_id}")
    assert resp_del.status_code == 404
    assert resp_del.is_json
    assert resp_del.json["success"] is False


def test_api_deletar_conta_logada(client):
    _criar_usuario_e_logar(client)
    resp = client.delete("/api/account", json={"confirmacao": "EXCLUIR"})
    assert resp.status_code == 200
    assert resp.is_json
    assert resp.json["success"] is True

    # Deve estar deslogado
    resp2 = client.get("/api/profile")
    assert resp2.status_code == 401


def test_dashboard_sem_registros_mostra_mensagem_graficos(client):
    _criar_usuario_e_logar(client)
    response = client.get("/dashboard")
    assert response.status_code == 200
    assert "Ainda não há dados suficientes para gerar o gráfico.".encode("utf-8") in response.data


def test_dashboard_graficos_apenas_do_usuario_logado(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    # Insere registro finalizado para u1
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, "2026-01-01", "08:00", "12:00", "13:00", "17:00"),
    )
    conn.commit()
    conn.close()

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")

    # Insere registro finalizado para u2
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (2, "2026-02-02", "08:00", "12:00", "13:00", "18:00"),
    )
    conn.commit()
    conn.close()

    response = client.get("/dashboard")
    assert response.status_code == 200
    # Deve conter apenas a data do usuário logado (u2)
    assert b"2026-02-02" in response.data
    assert b"2026-01-01" not in response.data


def test_export_excel_sem_login_falha(client):
    resp = client.get("/export/excel", follow_redirects=False)
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/")


def test_export_excel_retorna_arquivo_valido_e_so_do_usuario(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, "2026-03-10", "08:00", "12:00", "13:00", "17:00"),
    )
    conn.commit()
    conn.close()

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (2, "2026-04-11", "08:00", "12:00", "13:00", "18:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/export/excel")
    assert resp.status_code == 200
    assert resp.headers["Content-Disposition"].startswith("attachment")
    assert "registros_ponto.xlsx" in resp.headers["Content-Disposition"]

    wb = openpyxl.load_workbook(BytesIO(resp.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0].startswith("Período:")
    assert rows[2] == (
        "Data",
        "Entrada",
        "Saída Almoço",
        "Volta Almoço",
        "Saída Final",
        "Total Trabalhado",
        "Saldo do Dia",
    )

    # Deve conter apenas registro do u2
    valores = "\n".join(str(c) for r in rows[3:] for c in r if c is not None)
    assert "2026-04-11" in valores
    assert "2026-03-10" not in valores


def test_dashboard_filtro_periodo_retorna_apenas_registros_do_intervalo(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.executemany(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            (1, "2026-01-01", "08:00", "12:00", "13:00", "17:00"),
            (1, "2026-02-01", "08:00", "12:00", "13:00", "17:00"),
            (1, "2026-03-01", "08:00", "12:00", "13:00", "17:00"),
        ],
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard?data_inicio=2026-02-01&data_fim=2026-02-28")
    assert resp.status_code == 200
    assert b"2026-02-01" in resp.data
    assert b"2026-01-01" not in resp.data
    assert b"2026-03-01" not in resp.data


def test_export_excel_respeita_filtro_periodo(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.executemany(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            (1, "2026-01-01", "08:00", "12:00", "13:00", "17:00"),
            (1, "2026-02-15", "08:00", "12:00", "13:00", "17:00"),
        ],
    )
    conn.commit()
    conn.close()

    resp = client.get("/export/excel?data_inicio=2026-02-01&data_fim=2026-02-28")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(resp.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Período: 01/02/2026 até 28/02/2026"

    valores = "\n".join(str(c) for r in rows[3:] for c in r if c is not None)
    assert "2026-02-15" in valores
    assert "2026-01-01" not in valores


def test_dashboard_valida_periodo_invalido(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    resp = client.get("/dashboard?data_inicio=2026-03-10&data_fim=2026-03-01", follow_redirects=False)
    assert resp.status_code == 302
    assert resp.headers["Location"].startswith("/dashboard")


def test_export_pdf_sem_login_falha(client):
    resp = client.get("/export/pdf", follow_redirects=False)
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/")


def test_export_pdf_retorna_pdf_valido_e_so_do_usuario(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, "2026-03-10", "08:00", "12:00", "13:00", "17:00"),
    )
    conn.commit()
    conn.close()

    client.get("/logout")
    _criar_usuario_e_logar(client, "u2", "1234")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (2, "2026-04-11", "08:00", "12:00", "13:00", "18:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/export/pdf")
    assert resp.status_code == 200
    assert resp.headers["Content-Disposition"].startswith("attachment")
    assert "registros_ponto.pdf" in resp.headers["Content-Disposition"]
    assert resp.mimetype == "application/pdf"
    assert resp.data.startswith(b"%PDF")

    # Como o PDF é gerado sem compressão, o texto deve estar presente.
    assert b"Relat" in resp.data  # "Relatório de Ponto" pode variar por acentuação/fonte no PDF
    assert b"2026-04-11" in resp.data
    assert b"2026-03-10" not in resp.data


def test_export_pdf_respeita_filtro_periodo(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.executemany(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            (1, "2026-01-01", "08:00", "12:00", "13:00", "17:00"),
            (1, "2026-02-15", "08:00", "12:00", "13:00", "17:00"),
        ],
    )
    conn.commit()
    conn.close()

    resp = client.get("/export/pdf?data_inicio=2026-02-01&data_fim=2026-02-28")
    assert resp.status_code == 200
    assert resp.data.startswith(b"%PDF")
    assert b"01/02/2026" in resp.data
    assert b"28/02/2026" in resp.data
    assert b"2026-02-15" in resp.data
    assert b"2026-01-01" not in resp.data


def test_export_pdf_nao_quebra_com_muitos_registros(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    rows = []
    for i in range(1, 41):
        data = f"2026-03-{i:02d}" if i <= 28 else f"2026-04-{(i-28):02d}"
        rows.append((1, data, "08:00", "12:00", "13:00", "17:00"))
    cursor.executemany(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    conn.commit()
    conn.close()

    resp = client.get("/export/pdf")
    assert resp.status_code == 200
    assert resp.data.startswith(b"%PDF")


def test_alterar_senha_exige_senha_atual_correta(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    resp = client.post(
        "/perfil/alterar-senha",
        data={"senha_atual": "errada", "nova_senha": "nova", "confirmar_nova_senha": "nova"},
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert "Senha atual incorreta.".encode("utf-8") in resp.data


def test_alterar_senha_sucesso_e_login_com_nova_senha(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    resp = client.post(
        "/perfil/alterar-senha",
        data={"senha_atual": "1234", "nova_senha": "nova123", "confirmar_nova_senha": "nova123"},
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert "Senha alterada com sucesso.".encode("utf-8") in resp.data

    client.get("/logout")
    resp_login = client.post("/", data={"username": "u1", "password": "nova123"}, follow_redirects=False)
    assert resp_login.status_code == 302
    assert "/dashboard" in resp_login.headers["Location"]


def test_dashboard_mostra_botao_este_mes_e_banco_de_horas(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    # Define esperado: 8h (default já é 8h, mas mantém explícito)
    client.post("/perfil", data={"horas_diarias_esperadas": "8"}, follow_redirects=True)

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    # +1h
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, "2026-04-10", "08:00", "12:00", "13:00", "18:00"),
    )
    # -1h
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, "2026-04-11", "08:00", "12:00", "13:00", "16:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert b"Este m\xc3\xaas" in resp.data
    assert "Banco de horas do período:".encode("utf-8") in resp.data
    assert b"+0h 00m" in resp.data
    assert "Banco de horas do mês atual:".encode("utf-8") in resp.data


def test_dashboard_banco_mes_atual_considera_apenas_mes_corrente_e_finalizados(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    from datetime import date

    hoje = date.today()
    mes_inicio = hoje.replace(day=1)
    dia_mes = mes_inicio.replace(day=min(2, 28))
    # dia fora do mês (mês anterior)
    if mes_inicio.month == 1:
        fora_mes = mes_inicio.replace(year=mes_inicio.year - 1, month=12, day=28)
    else:
        fora_mes = mes_inicio.replace(month=mes_inicio.month - 1, day=28)

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    # Dentro do mês: +1h
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, dia_mes.isoformat(), "08:00", "12:00", "13:00", "18:00"),
    )
    # Dentro do mês: em aberto (não conta)
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha)
        VALUES (?, ?, ?)
        """,
        (1, mes_inicio.isoformat(), "08:00"),
    )
    # Fora do mês: +2h (não conta)
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, fora_mes.isoformat(), "08:00", "12:00", "13:00", "19:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Banco de horas do mês atual: +1h 00m".encode("utf-8") in resp.data


def test_404_amigavel_mostra_botao_voltar(client):
    resp = client.get("/nao-existe")
    assert resp.status_code == 404
    assert b"404" in resp.data

    _criar_usuario_e_logar(client, "u1", "1234")
    resp2 = client.get("/nao-existe-2")
    assert resp2.status_code == 404
    assert b"/dashboard" in resp2.data


def test_status_dia_nao_iniciada(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Não iniciada".encode("utf-8") in resp.data


def test_status_dia_em_andamento(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    # cria entrada hoje em aberto via rota existente
    client.get("/bater")
    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Em andamento".encode("utf-8") in resp.data


def test_status_dia_jornada_completa(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    client.post("/perfil", data={"horas_diarias_esperadas": "8"}, follow_redirects=True)

    hoje = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, hoje, "08:00", "12:00", "13:00", "17:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Jornada completa".encode("utf-8") in resp.data


def test_status_dia_jornada_incompleta(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    client.post("/perfil", data={"horas_diarias_esperadas": "8"}, follow_redirects=True)

    hoje = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, hoje, "08:00", "12:00", "13:00", "16:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Jornada incompleta".encode("utf-8") in resp.data


def test_tabela_saldo_recebe_classe_por_sinal(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    client.post("/perfil", data={"horas_diarias_esperadas": "8"}, follow_redirects=True)

    hoje = datetime.now().strftime("%Y-%m-%d")
    ontem = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    anteontem = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    # +1h
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, anteontem, "08:00", "12:00", "13:00", "18:00"),
    )
    # -1h
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, ontem, "08:00", "12:00", "13:00", "16:00"),
    )
    # 0h
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, hoje, "08:00", "12:00", "13:00", "17:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert b"class=\"text-success\">+1h 00m" in resp.data
    assert b"class=\"text-danger\">-1h 00m" in resp.data
    assert b"class=\"text-secondary\">+0h 00m" in resp.data


def test_resumo_mensal_sem_registros_no_mes(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Resumo do mês".encode("utf-8") in resp.data
    assert b"Dias trabalhados" in resp.data
    assert b">0<" in resp.data  # dias


def test_resumo_mensal_calculos_com_multiplos_dias(client):
    _criar_usuario_e_logar(client, "u1", "1234")

    from datetime import date

    hoje = date.today()
    mes_inicio = hoje.replace(day=1).isoformat()
    dia2 = hoje.replace(day=min(2, 28)).isoformat()

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.executemany(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            # 9h -> 9h 00m
            (1, mes_inicio, "08:00", "12:00", "13:00", "18:00"),
            # 7h -> 7h 00m
            (1, dia2, "08:00", "12:00", "13:00", "16:00"),
        ],
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    # Dias trabalhados: 2
    assert b">2<" in resp.data
    # Total: 16h 00m
    assert "16h 00m".encode("utf-8") in resp.data
    # Média: 8h 00m
    assert "8h 00m".encode("utf-8") in resp.data


def test_botao_ponto_sem_registro_hoje_mostra_registrar_entrada(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Registrar entrada".encode("utf-8") in resp.data


def test_botao_ponto_com_entrada_sem_saida_almoco_mostra_saida_almoco(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    hoje = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO registros (user_id, data, entrada_manha) VALUES (?, ?, ?)",
        (1, hoje, "08:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Registrar saída almoço".encode("utf-8") in resp.data


def test_botao_ponto_com_saida_almoco_sem_volta_mostra_volta_almoco(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    hoje = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco)
        VALUES (?, ?, ?, ?)
        """,
        (1, hoje, "08:00", "12:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Registrar volta almoço".encode("utf-8") in resp.data


def test_botao_ponto_com_volta_almoco_sem_saida_final_mostra_saida_final(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    hoje = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco)
        VALUES (?, ?, ?, ?, ?)
        """,
        (1, hoje, "08:00", "12:00", "13:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Registrar saída final".encode("utf-8") in resp.data


def test_botao_ponto_com_saida_final_mostra_registrar_novo_ponto(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    hoje = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO registros (user_id, data, entrada_manha, saida_almoco, volta_almoco, saida_final)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (1, hoje, "08:00", "12:00", "13:00", "17:00"),
    )
    conn.commit()
    conn.close()

    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "Registrar novo ponto".encode("utf-8") in resp.data


def test_deletar_conta_apaga_registros_e_desloga(client):
    _criar_usuario_e_logar(client)
    client.get("/bater")

    response = client.post("/perfil/excluir-conta", data={"confirmacao": "EXCLUIR"}, follow_redirects=False)
    assert response.status_code == 302
    assert response.headers["Location"].endswith("/")

    # Acessar dashboard deve redirecionar (sessão encerrada)
    response_dashboard = client.get("/dashboard", follow_redirects=False)
    assert response_dashboard.status_code == 302
    assert response_dashboard.headers["Location"].endswith("/")

    conn = sqlite3.connect("web/database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM usuarios")
    usuarios = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM registros")
    registros = cursor.fetchone()[0]
    conn.close()

    assert usuarios == 0
    assert registros == 0


def test_flash_exclusao_conta_nao_persiste_para_outro_usuario(client):
    _criar_usuario_e_logar(client, "u1", "1234")
    response = client.post("/perfil/excluir-conta", data={"confirmacao": "EXCLUIR"}, follow_redirects=True)
    assert response.status_code == 200
    assert "Conta excluída com sucesso.".encode("utf-8") in response.data

    # Criar e logar outro usuário não deve ver o flash antigo no dashboard
    _criar_usuario_e_logar(client, "u2", "1234")
    response_dashboard = client.get("/dashboard")
    assert response_dashboard.status_code == 200
    assert "Conta excluída com sucesso.".encode("utf-8") not in response_dashboard.data

